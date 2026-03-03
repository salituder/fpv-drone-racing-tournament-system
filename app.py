import sqlite3
import random
import math
import os
import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st

DB_PATH = "tournament.db"

# ============================================================
# БАЗОВЫЕ СТИЛИ CSS
# ============================================================

BASE_CSS = """
<style>
.tournament-progress {
    display: flex;
    align-items: flex-start;
    margin: 20px 0;
    position: relative;
    padding: 0 10px;
}
.progress-step {
    display: flex;
    flex-direction: column;
    align-items: center;
    flex: 1;
    position: relative;
    z-index: 1;
}
.progress-step:not(:last-child)::after {
    content: '';
    position: absolute;
    top: 14px;
    left: 50%;
    width: 100%;
    height: 3px;
    background: #3a3a3a;
    z-index: 0;
}
.progress-step.completed:not(:last-child)::after {
    background: #4CAF50;
}
.progress-step.active:not(:last-child)::after {
    background: linear-gradient(90deg, #667eea 0%, #3a3a3a 100%);
}
.progress-dot {
    width: 28px;
    height: 28px;
    border-radius: 50%;
    background: #3a3a3a;
    border: 3px solid #555;
    z-index: 2;
    position: relative;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 12px;
}
.progress-step.completed .progress-dot {
    background: #4CAF50;
    border-color: #45a049;
    box-shadow: 0 0 8px rgba(76, 175, 80, 0.5);
}
.progress-step.active .progress-dot {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    border-color: #667eea;
    box-shadow: 0 0 12px rgba(102, 126, 234, 0.6);
    animation: pulse-dot 2s infinite;
}
@keyframes pulse-dot {
    0%, 100% { box-shadow: 0 0 8px rgba(102, 126, 234, 0.4); }
    50% { box-shadow: 0 0 16px rgba(102, 126, 234, 0.8); }
}
.progress-step.pending .progress-dot {
    background: #2a2a2a;
    border-color: #444;
}
.progress-label {
    margin-top: 8px;
    font-size: 0.75em;
    font-weight: 500;
    text-align: center;
    max-width: 90px;
    line-height: 1.2;
    color: #888;
}
.progress-step.completed .progress-label {
    color: #4CAF50;
}
.progress-step.active .progress-label {
    color: #667eea;
    font-weight: 700;
}
/* Bracket tree */
.bracket-container {
    display: flex;
    align-items: center;
    gap: 0;
    overflow-x: auto;
    padding: 20px 0;
}
.bracket-round-wrapper {
    display: flex;
    flex-direction: column;
    align-items: stretch;
    flex-shrink: 0;
}
.bracket-round-title {
    text-align: center;
    font-weight: 700;
    font-size: 1em;
    margin-bottom: 8px;
    padding: 6px 12px;
    border-radius: 8px;
    background: #2a2a2a;
    color: #ccc;
}
.bracket-round-title.active-round {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
}
.bracket-round-title.done-round {
    background: #2e7d32;
    color: #90EE90;
}
.bracket-round-title.final-round {
    background: linear-gradient(135deg, #d4a017 0%, #b8860b 100%);
    color: white;
}
.bracket-groups-row {
    display: flex;
    align-items: stretch;
}
.bracket-groups-col {
    display: flex;
    flex-direction: column;
    justify-content: center;
    min-width: 280px;
    max-width: 340px;
}
.bracket-group {
    background: #1e1e1e;
    border: 1px solid #444;
    border-radius: 8px;
    margin: 6px 0;
    padding: 8px 10px;
}
.bracket-group-title {
    font-size: 0.78em;
    color: #888;
    margin-bottom: 4px;
    text-align: center;
    font-weight: 600;
}
.bracket-player {
    display: flex;
    justify-content: space-between;
    padding: 4px 8px;
    font-size: 0.85em;
    border-radius: 4px;
    margin: 2px 0;
    gap: 12px;
}
.bracket-player.advancing {
    background: #1a3a1a;
    color: #90EE90;
}
.bracket-player.eliminated {
    background: #3a1a1a;
    color: #FFB6B6;
}
.bracket-player.gold {
    background: #5C4B00;
    color: #FFD700;
    font-weight: 700;
}
.bracket-player.silver {
    background: #3A3A3A;
    color: #C0C0C0;
}
.bracket-player.bronze {
    background: #3D2B1F;
    color: #CD7F32;
}
.bracket-player.pending-player {
    color: #666;
}
.bracket-connector {
    display: flex;
    flex-direction: column;
    justify-content: center;
    width: 60px;
    min-width: 60px;
    flex-shrink: 0;
    margin-right: 10px;
}
.bracket-conn-top {
    flex: 1;
    border-right: 2px solid #555;
    border-top: 2px solid #555;
    margin-left: 40%;
    min-height: 10px;
}
.bracket-conn-bottom {
    flex: 1;
    border-right: 2px solid #555;
    border-bottom: 2px solid #555;
    margin-left: 40%;
    min-height: 10px;
}
</style>
"""

# ============================================================
# i18n — Русский по умолчанию
# ============================================================

I18N = {
    "RU": {
        "app_title": "🏁 Турнир по гонкам дронов (беспилотных воздушных судов)",
        "language": "Язык",
        "tournament": "Турнир",
        "select_tournament": "Выбрать турнир",
        "create_new": "➕ Создать новый",
        "create_new_header": "Новый турнир",
        "tournament_name": "Название турнира",
        "discipline": "Дисциплина",
        "create_tournament": "Создать турнир",
        "time_limit": "Лимит времени (сек)",
        "total_laps": "Кол-во кругов (по регламенту)",

        # Дисциплины
        "drone_individual": "Дроны: Личный зачёт",
        "sim_individual": "Симулятор: Личный зачёт",
        "sim_team": "Симулятор: Командный зачёт",
        "coming_soon": "В разработке...",

        # Навигация
        "nav_overview": "📊 Обзор",
        "nav_participants": "👥 Участники",
        "nav_qualification": "⏱️ Квалификация",
        "nav_bracket": "🏆 Сетка",
        "nav_playoff": "🔥 Плей-офф",
        "nav_final": "🥇 Финал",

        # Обзор
        "overview_title": "Обзор турнира",
        "total_participants": "Участников",
        "tournament_status": "Статус",
        "status_setup": "Подготовка",
        "status_qualification": "Квалификация",
        "status_bracket": "Плей-офф",
        "status_finished": "Завершён",

        # Участники
        "participants_title": "Список участников",
        "add_participant": "Добавить участника",
        "pilot_name": "Имя пилота",
        "add": "Добавить",
        "random_draw": "🎲 Жеребьёвка (случайные номера)",
        "draw_done": "Жеребьёвка проведена!",
        "draw_already": "Жеребьёвка уже проведена",
        "start_number": "Стартовый №",
        "demo_fill": "Тестовое заполнение",
        "demo_hint": "Быстро добавить тестовых участников",
        "demo_count": "Количество",
        "demo_prefix": "Префикс имени",
        "demo_add": "Добавить тестовых",
        "demo_already": "Участники уже добавлены",
        "demo_added": "Добавлено участников",

        # Квалификация
        "qual_title": "Квалификационный этап",
        "qual_info": "Введите результаты каждого пилота. Система автоматически определит кто проходит дальше.",
        "time_seconds": "Время (сек)",
        "laps_completed": "Круги.Препятствия",
        "completed_all": "Все 3 круга",
        "projected_time": "Расчётное время (3 кр.)",
        "qual_rank": "Место",
        "qual_cutoff": "Проходят: {} из {}",
        "qual_finish": "✅ Завершить квалификацию → Сформировать сетку",
        "qual_not_all": "Не все результаты введены!",
        "qual_done": "Квалификация завершена!",

        # Сетка
        "bracket_title": "Турнирная сетка",
        "group": "Группа",
        "advance_stage": "➡️ Перейти к следующему этапу",
        "start_playoff": "🚀 Начать плей-офф",
        "last_stage": "Это финальный этап",
        "tie_warning": "⚠️ Обнаружено равенство очков! Возможно потребуется доп. вылет.",
        "waiting_for_qual": "Ожидание завершения квалификации",

        # Плей-офф
        "playoff_title": "Плей-офф — ввод результатов",
        "playoff_not_started": "Плей-офф ещё не начался",
        "select_round": "Выберите раунд",
        "select_group": "Группа",

        # Финал
        "final_title": "ФИНАЛ",
        "heat_n": "Вылет {}",
        "final_standings": "Итоговая таблица финала",
        "champion": "🏆 ЧЕМПИОН",
        "bonus_note": "Бонус +1 за 2 и более побед в вылетах",

        # Симулятор
        "track": "Трасса",
        "track_n": "Трасса {}",
        "attempt_n": "Попытка {}",
        "sim_qual_info": "Введите результаты каждого пилота. Лимит 2 минуты, 3 круга.",
        "sim_group_results": "Сводка группы",

        # Общее
        "saved": "✅ Сохранено!",
        "error": "Ошибка",
        "download_csv": "📥 Скачать CSV",
        "place_short": "м.",
        "points": "Очки",
        "pilot": "Пилот",
        "time": "Время",
        "laps": "Круги",
        "place": "Место",

        # Дисквалификация
        "disqualify": "Дисквалифицировать (техпор)",
        "disqualify_undo": "Снять дисквалификацию",
        "disqualified": "DSQ",
        "disqualified_full": "Дисквалифицирован",
    },
    "EN": {
        "app_title": "🏁 UAV Drone Racing Tournament",
        "language": "Language",
        "tournament": "Tournament",
        "select_tournament": "Select tournament",
        "create_new": "➕ Create new",
        "create_new_header": "New tournament",
        "tournament_name": "Tournament name",
        "discipline": "Discipline",
        "create_tournament": "Create tournament",
        "time_limit": "Time limit (sec)",
        "total_laps": "Laps count (regulation)",

        "drone_individual": "Drones: Individual",
        "sim_individual": "Simulator: Individual",
        "sim_team": "Simulator: Team",
        "coming_soon": "Coming soon...",

        "nav_overview": "📊 Overview",
        "nav_participants": "👥 Participants",
        "nav_qualification": "⏱️ Qualification",
        "nav_bracket": "🏆 Bracket",
        "nav_playoff": "🔥 Playoff",
        "nav_final": "🥇 Final",

        "overview_title": "Tournament Overview",
        "total_participants": "Participants",
        "tournament_status": "Status",
        "status_setup": "Setup",
        "status_qualification": "Qualification",
        "status_bracket": "Playoff",
        "status_finished": "Finished",

        "participants_title": "Participants",
        "add_participant": "Add participant",
        "pilot_name": "Pilot name",
        "add": "Add",
        "random_draw": "🎲 Random draw",
        "draw_done": "Draw completed!",
        "draw_already": "Draw already done",
        "start_number": "Start #",
        "demo_fill": "Test fill",
        "demo_hint": "Quick add test participants",
        "demo_count": "Count",
        "demo_prefix": "Name prefix",
        "demo_add": "Add test",
        "demo_already": "Participants already added",
        "demo_added": "Added participants",

        "qual_title": "Qualification",
        "qual_info": "Enter results for each pilot. System auto-determines who advances.",
        "time_seconds": "Time (sec)",
        "laps_completed": "Laps.Obstacles",
        "completed_all": "All 3 laps",
        "projected_time": "Projected time (3 laps)",
        "qual_rank": "Rank",
        "qual_cutoff": "Advancing: {} of {}",
        "qual_finish": "✅ Finish Qualification → Generate Bracket",
        "qual_not_all": "Not all results entered!",
        "qual_done": "Qualification complete!",

        "bracket_title": "Tournament Bracket",
        "group": "Group",
        "advance_stage": "➡️ Advance to next stage",
        "start_playoff": "🚀 Start Playoff",
        "last_stage": "This is the final stage",
        "tie_warning": "⚠️ Tie detected! Extra heat may be required.",
        "waiting_for_qual": "Waiting for qualification",

        "playoff_title": "Playoff — Enter Results",
        "playoff_not_started": "Playoff not started yet",
        "select_round": "Select round",
        "select_group": "Group",

        "final_title": "FINAL",
        "heat_n": "Heat {}",
        "final_standings": "Final Standings",
        "champion": "🏆 CHAMPION",
        "bonus_note": "Bonus +1 for 2+ first-place finishes",

        "track": "Track",
        "track_n": "Track {}",
        "attempt_n": "Attempt {}",
        "sim_qual_info": "Enter results for each pilot. 2 minute limit, 3 laps.",
        "sim_group_results": "Group summary",

        "saved": "✅ Saved!",
        "error": "Error",
        "download_csv": "📥 Download CSV",
        "place_short": "pl.",
        "points": "Points",
        "pilot": "Pilot",
        "time": "Time",
        "laps": "Laps",
        "place": "Place",

        "disqualify": "Disqualify (technical defeat)",
        "disqualify_undo": "Remove disqualification",
        "disqualified": "DSQ",
        "disqualified_full": "Disqualified",
    }
}


def T(key: str) -> str:
    lang = st.session_state.get("lang", "RU")
    return I18N.get(lang, I18N["RU"]).get(key, I18N["RU"].get(key, key))


# ============================================================
# ТАБЛИЦЫ ПОСЕВА И ПРОГРЕССА (из официального регламента)
# ============================================================

# Посев 32 → 1/8 (Таблица №3)
SEEDING_1_8_32: Dict[int, List[int]] = {
    1: [1, 9, 24, 32], 2: [8, 16, 17, 25], 3: [7, 15, 18, 26], 4: [6, 14, 19, 27],
    5: [5, 13, 20, 28], 6: [4, 12, 21, 29], 7: [3, 11, 22, 30], 8: [2, 10, 23, 31],
}

# Посев 16 → 1/4 (Таблица №4)
SEEDING_1_4_16: Dict[int, List[int]] = {
    1: [1, 5, 12, 16], 2: [3, 7, 10, 14], 3: [2, 6, 11, 15], 4: [4, 8, 9, 13],
}

# Посев 8 → 1/2 (аналогичная схема змейкой)
SEEDING_1_2_8: Dict[int, List[int]] = {
    1: [1, 4, 5, 8], 2: [2, 3, 6, 7],
}

# Посев 4 → Финал
SEEDING_FINAL_4: Dict[int, List[int]] = {
    1: [1, 2, 3, 4],
}

# Пересев 1/8 → 1/4
PROGRESS_1_8_TO_1_4: Dict[int, List[Tuple[int, int]]] = {
    1: [(1, 1), (1, 5), (2, 6), (2, 2)],
    2: [(1, 7), (1, 3), (2, 8), (2, 4)],
    3: [(1, 8), (1, 4), (2, 7), (2, 3)],
    4: [(1, 6), (1, 2), (2, 1), (2, 5)],
}

# Пересев 1/4 → 1/2
PROGRESS_1_4_TO_1_2: Dict[int, List[Tuple[int, int]]] = {
    1: [(1, 1), (1, 2), (2, 3), (2, 4)],
    2: [(1, 3), (1, 4), (2, 1), (2, 2)],
}

# Пересев 1/2 → Финал
PROGRESS_1_2_TO_FINAL: Dict[int, List[Tuple[int, int]]] = {
    1: [(1, 1), (1, 2), (2, 1), (2, 2)]
}

# Очки финала (дроны)
FINAL_SCORING = {1: 3, 2: 2, 3: 1, 4: 0}

# Очки группового/финального этапа (симулятор)
# 4 пилота летят одновременно, 2 трассы × 3 попытки = 6 вылетов, сумма очков. Макс 24.
SIM_SCORING = {1: 4, 2: 3, 3: 2, 4: 1}  # 0 для DNF (не в словаре)


# ============================================================
# StageDef + динамическая генерация сетки
# ============================================================

@dataclass
class StageDef:
    code: str
    display_name: Dict[str, str]
    group_size: int
    group_count: int
    qualifiers: int
    heats_count: int = 1  # 3 для финала
    seeding_map: Optional[Dict[int, List[int]]] = None
    progress_map: Optional[Dict[int, List[Tuple[int, int]]]] = None


def compute_bracket_size(n: int) -> int:
    """Наибольшая степень 2 <= n (минимум 4)."""
    for s in [32, 16, 8, 4]:
        if n >= s:
            return s
    return 4


def generate_bracket(advancing: int) -> List[StageDef]:
    """Генерирует список этапов плей-офф по количеству прошедших."""
    stages: List[StageDef] = []
    if advancing >= 32:
        stages.append(StageDef("1/8", {"RU": "1/8 финала", "EN": "Round of 16"}, 4, 8, 2, 1,
                                seeding_map=SEEDING_1_8_32))
        stages.append(StageDef("1/4", {"RU": "Четвертьфинал", "EN": "Quarterfinal"}, 4, 4, 2, 1,
                                progress_map=PROGRESS_1_8_TO_1_4))
        stages.append(StageDef("1/2", {"RU": "Полуфинал", "EN": "Semifinal"}, 4, 2, 2, 1,
                                progress_map=PROGRESS_1_4_TO_1_2))
        stages.append(StageDef("F", {"RU": "ФИНАЛ", "EN": "FINAL"}, 4, 1, 0, 3,
                                progress_map=PROGRESS_1_2_TO_FINAL))
    elif advancing >= 16:
        stages.append(StageDef("1/4", {"RU": "Четвертьфинал", "EN": "Quarterfinal"}, 4, 4, 2, 1,
                                seeding_map=SEEDING_1_4_16))
        stages.append(StageDef("1/2", {"RU": "Полуфинал", "EN": "Semifinal"}, 4, 2, 2, 1,
                                progress_map=PROGRESS_1_4_TO_1_2))
        stages.append(StageDef("F", {"RU": "ФИНАЛ", "EN": "FINAL"}, 4, 1, 0, 3,
                                progress_map=PROGRESS_1_2_TO_FINAL))
    elif advancing >= 8:
        stages.append(StageDef("1/2", {"RU": "Полуфинал", "EN": "Semifinal"}, 4, 2, 2, 1,
                                seeding_map=SEEDING_1_2_8))
        stages.append(StageDef("F", {"RU": "ФИНАЛ", "EN": "FINAL"}, 4, 1, 0, 3,
                                progress_map=PROGRESS_1_2_TO_FINAL))
    else:  # 4
        stages.append(StageDef("F", {"RU": "ФИНАЛ", "EN": "FINAL"}, 4, 1, 0, 3,
                                seeding_map=SEEDING_FINAL_4))
    return stages


# ============================================================
# База данных
# ============================================================

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = db()
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS tournaments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        discipline TEXT NOT NULL DEFAULT 'drone_individual',
        time_limit_seconds REAL NOT NULL DEFAULT 90.0,
        total_laps INTEGER NOT NULL DEFAULT 3,
        scoring_mode TEXT NOT NULL DEFAULT 'none',
        status TEXT NOT NULL DEFAULT 'setup',
        created_at TEXT NOT NULL
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS participants(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tournament_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        start_number INTEGER,
        FOREIGN KEY(tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS qualification_results(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tournament_id INTEGER NOT NULL,
        participant_id INTEGER NOT NULL,
        time_seconds REAL,
        laps_completed REAL,
        completed_all_laps INTEGER DEFAULT 0,
        projected_time REAL,
        UNIQUE(tournament_id, participant_id),
        FOREIGN KEY(tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE,
        FOREIGN KEY(participant_id) REFERENCES participants(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS stages(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tournament_id INTEGER NOT NULL,
        stage_idx INTEGER NOT NULL,
        code TEXT NOT NULL,
        group_size INTEGER NOT NULL,
        group_count INTEGER NOT NULL,
        qualifiers INTEGER NOT NULL,
        heats_count INTEGER NOT NULL DEFAULT 1,
        status TEXT NOT NULL DEFAULT 'active',
        UNIQUE(tournament_id, stage_idx),
        FOREIGN KEY(tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS groups(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stage_id INTEGER NOT NULL,
        group_no INTEGER NOT NULL,
        UNIQUE(stage_id, group_no),
        FOREIGN KEY(stage_id) REFERENCES stages(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS group_members(
        group_id INTEGER NOT NULL,
        participant_id INTEGER NOT NULL,
        PRIMARY KEY(group_id, participant_id),
        FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE,
        FOREIGN KEY(participant_id) REFERENCES participants(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS heats(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        group_id INTEGER NOT NULL,
        heat_no INTEGER NOT NULL,
        track_no INTEGER NOT NULL DEFAULT 1,
        UNIQUE(group_id, track_no, heat_no),
        FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS heat_results(
        heat_id INTEGER NOT NULL,
        participant_id INTEGER NOT NULL,
        time_seconds REAL,
        laps_completed REAL,
        completed_all_laps INTEGER DEFAULT 0,
        projected_time REAL,
        place INTEGER,
        points INTEGER DEFAULT 0,
        PRIMARY KEY(heat_id, participant_id),
        FOREIGN KEY(heat_id) REFERENCES heats(id) ON DELETE CASCADE,
        FOREIGN KEY(participant_id) REFERENCES participants(id) ON DELETE CASCADE
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS team_pilots(
        participant_id INTEGER PRIMARY KEY,
        pilot1_name TEXT NOT NULL,
        pilot2_name TEXT NOT NULL,
        FOREIGN KEY(participant_id) REFERENCES participants(id) ON DELETE CASCADE
    )""")

    # === Миграции: добавить столбцы, если их нет (для обновления старых БД) ===
    try:
        c.execute("ALTER TABLE tournaments ADD COLUMN scoring_mode TEXT NOT NULL DEFAULT 'none'")
    except sqlite3.OperationalError:
        pass  # столбец уже существует
    try:
        c.execute("ALTER TABLE participants ADD COLUMN disqualified INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass  # столбец уже существует
    try:
        c.execute("ALTER TABLE tournaments ADD COLUMN qual_attempts INTEGER NOT NULL DEFAULT 1")
    except sqlite3.OperationalError:
        pass  # столбец уже существует

    # Миграция qualification_results: добавить attempt_no для нескольких попыток
    qr_cols = [row[1] for row in c.execute("PRAGMA table_info(qualification_results)").fetchall()]
    if "attempt_no" not in qr_cols:
        c.execute("""CREATE TABLE qualification_results_new(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tournament_id INTEGER NOT NULL,
            participant_id INTEGER NOT NULL,
            attempt_no INTEGER NOT NULL DEFAULT 1,
            time_seconds REAL,
            laps_completed REAL,
            completed_all_laps INTEGER DEFAULT 0,
            projected_time REAL,
            UNIQUE(tournament_id, participant_id, attempt_no),
            FOREIGN KEY(tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE,
            FOREIGN KEY(participant_id) REFERENCES participants(id) ON DELETE CASCADE
        )""")
        c.execute("""INSERT INTO qualification_results_new(id, tournament_id, participant_id, attempt_no,
            time_seconds, laps_completed, completed_all_laps, projected_time)
            SELECT id, tournament_id, participant_id, 1, time_seconds, laps_completed, completed_all_laps, projected_time
            FROM qualification_results""")
        c.execute("DROP TABLE qualification_results")
        c.execute("ALTER TABLE qualification_results_new RENAME TO qualification_results")

    # Миграция heats: пересоздаём таблицу с правильным UNIQUE constraint
    # Проверяем, есть ли столбец track_no и правильный ли constraint
    cols = [row[1] for row in c.execute("PRAGMA table_info(heats)").fetchall()]
    if "track_no" not in cols:
        # Старая таблица без track_no — пересоздаём
        c.execute("""CREATE TABLE IF NOT EXISTS heats_new(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id INTEGER NOT NULL,
            heat_no INTEGER NOT NULL,
            track_no INTEGER NOT NULL DEFAULT 1,
            UNIQUE(group_id, track_no, heat_no),
            FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
        )""")
        c.execute("INSERT INTO heats_new(id, group_id, heat_no, track_no) SELECT id, group_id, heat_no, 1 FROM heats")
        c.execute("DROP TABLE heats")
        c.execute("ALTER TABLE heats_new RENAME TO heats")
    else:
        # track_no есть, но возможно constraint старый — проверяем
        # Пробуем вставить тестовую запись, которая бы нарушила старый constraint
        # Безопаснее: пересоздать если autoindex всё ещё старый
        index_info = c.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='heats'").fetchone()
        if index_info and "UNIQUE(group_id, heat_no)" in (index_info[0] or ""):
            c.execute("""CREATE TABLE IF NOT EXISTS heats_new(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id INTEGER NOT NULL,
                heat_no INTEGER NOT NULL,
                track_no INTEGER NOT NULL DEFAULT 1,
                UNIQUE(group_id, track_no, heat_no),
                FOREIGN KEY(group_id) REFERENCES groups(id) ON DELETE CASCADE
            )""")
            c.execute("INSERT INTO heats_new(id, group_id, heat_no, track_no) SELECT id, group_id, heat_no, track_no FROM heats")
            c.execute("DROP TABLE heats")
            c.execute("ALTER TABLE heats_new RENAME TO heats")

    conn.commit()
    conn.close()


def qdf(sql, params=()) -> pd.DataFrame:
    conn = db()
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


def exec_sql(sql, params=()):
    conn = db()
    conn.execute(sql, params)
    conn.commit()
    conn.close()


def exec_many(sql, rows):
    conn = db()
    conn.executemany(sql, rows)
    conn.commit()
    conn.close()


def _detect_excel_discipline_columns(df: pd.DataFrame, scan_rows: int = 3) -> dict:
    """
    Сканирует заголовки Excel (строки 1–3): Дисциплина, под ней ТС/75, под ними ЛЗ.
    Находит колонки: ФИО, 75 ЛЗ (дроны), ТС ЛЗ (симулятор).
    """
    result = {}
    if df.empty or df.shape[1] == 0:
        return result
    n_cols = df.shape[1]
    n_rows = min(scan_rows, len(df))

    for col_idx in range(n_cols):
        parts = []
        for row_idx in range(n_rows):
            val = df.iloc[row_idx, col_idx]
            if pd.notna(val):
                s = str(val).strip()
                if s:
                    parts.append(s.lower())
        combined = "".join(parts).replace(" ", "").replace("\n", "")

        if "фио" in combined:
            result["fio"] = col_idx
        if "категория" in combined:
            result["category"] = col_idx
        if "75" in combined and "лз" in combined:
            result["75лз"] = col_idx
        if ("тс" in combined or "tc" in combined) and "лз" in combined:
            result["тслз"] = col_idx

    if "fio" not in result and n_cols > 1:
        result["fio"] = 1
    return result


def _parse_excel_discipline_list(df: pd.DataFrame, discipline: str,
                                  category_filter: Optional[str] = None, header_rows: int = 3) -> Tuple[List[str], dict]:
    """
    Парсит Excel: строки 1–3 — заголовки, с 4-й — данные.
    category_filter: "Мальчики"|"Юниорки"|"Юниоры"|"Девочки"|None (все)
    """
    cols = _detect_excel_discipline_columns(df, scan_rows=header_rows)
    disc_col = cols.get("75лз") if discipline == "drone_individual" else cols.get("тслз")
    fio_col = cols.get("fio", 1)
    cat_col = cols.get("category", 3)

    if disc_col is None:
        return [], cols

    names = []
    for row_idx in range(header_rows, len(df)):
        row = df.iloc[row_idx]
        if fio_col >= len(row):
            continue
        name = str(row.iloc[fio_col]).strip() if pd.notna(row.iloc[fio_col]) else ""
        if not name or name.lower() in ("nan", "none", ""):
            continue
        if disc_col < len(row):
            cell = row.iloc[disc_col]
            val = str(cell).strip() if pd.notna(cell) else ""
            if val != "+":
                continue
        if category_filter and cat_col < len(row):
            cat_val = str(row.iloc[cat_col]).strip() if pd.notna(row.iloc[cat_col]) else ""
            if cat_val != category_filter:
                continue
        names.append(name)
    return names, cols


# ============================================================
# Бизнес-логика: квалификация
# ============================================================

def calc_projected_time(time_seconds: float, laps_completed: float, total_laps: int = 3) -> Optional[float]:
    """Рассчитать теоретическое время на total_laps кругов."""
    if laps_completed and laps_completed > 0:
        return round(time_seconds * (total_laps / laps_completed), 2)
    return None


def _safe_time_for_input(seconds: float) -> float:
    """Для number_input(max_value=999): 9999 (DSQ sentinel), NaN и больше -> 0."""
    if seconds is None or (isinstance(seconds, float) and math.isnan(seconds)) or seconds > 999:
        return 0.0
    return float(seconds)


def rank_results(results: List[Dict]) -> List[Dict]:
    """
    Ранжирование по времени:
    1. Пролетели все круги → по time_seconds ASC
    2. Не долетели → по laps_completed DESC
    """
    completed = [r for r in results if r.get("completed_all_laps")]
    incomplete = [r for r in results if not r.get("completed_all_laps")]
    completed.sort(key=lambda r: r.get("time_seconds") or 9999)
    incomplete.sort(key=lambda r: -(r.get("laps_completed") or 0))
    ranked = completed + incomplete
    for i, r in enumerate(ranked):
        r["place"] = i + 1
    return ranked


def get_qualification_results(tournament_id: int) -> pd.DataFrame:
    """Получить результаты квалификации с ранжированием."""
    df = get_qual_ranking(tournament_id)
    return df


def get_participant_qual_attempts(tournament_id: int, participant_id: int) -> pd.DataFrame:
    """Получить все попытки квалификации участника."""
    return qdf("""
        SELECT attempt_no, time_seconds, laps_completed, completed_all_laps, projected_time
        FROM qualification_results
        WHERE tournament_id=? AND participant_id=? AND time_seconds IS NOT NULL
        ORDER BY attempt_no
    """, (tournament_id, participant_id))


def get_qual_ranking(tournament_id: int) -> pd.DataFrame:
    """Ранжированный список квалификации. При нескольких попытках — берётся лучший результат."""
    df = qdf("""
        SELECT p.id as pid, p.name, p.start_number,
               qr.attempt_no, qr.time_seconds, qr.laps_completed, qr.completed_all_laps, qr.projected_time,
               COALESCE(p.disqualified,0) as disqualified
        FROM participants p
        JOIN qualification_results qr ON qr.participant_id = p.id AND qr.tournament_id = ?
        WHERE p.tournament_id = ? AND qr.time_seconds IS NOT NULL
    """, (tournament_id, tournament_id))

    if df.empty:
        return df

    # При нескольких попытках — выбираем лучший результат по участнику
    best_rows = []
    for pid in df["pid"].unique():
        sub = df[df["pid"] == pid]
        recs = sub.to_dict("records")
        ranked_one = rank_results(recs)
        best_rows.append(ranked_one[0])
    df = pd.DataFrame(best_rows)

    results = df.to_dict("records")
    ranked = rank_results(results)
    ranked_df = pd.DataFrame(ranked)
    return _apply_dsq_to_ranking(ranked_df, tournament_id, "pid")


def save_qual_result(tournament_id: int, participant_id: int, time_seconds: float,
                     laps_completed: float, completed_all_laps: bool, total_laps: int = 3, attempt_no: int = 1):
    projected = calc_projected_time(time_seconds, laps_completed, total_laps) if not completed_all_laps else time_seconds
    exec_sql("""
        INSERT OR REPLACE INTO qualification_results(tournament_id, participant_id, attempt_no,
            time_seconds, laps_completed, completed_all_laps, projected_time)
        VALUES(?, ?, ?, ?, ?, ?, ?)
    """, (tournament_id, participant_id, attempt_no, time_seconds, laps_completed,
          int(completed_all_laps), projected))


def participant_count(tournament_id: int) -> int:
    df = qdf("SELECT COUNT(*) as c FROM participants WHERE tournament_id=?", (tournament_id,))
    return int(df.iloc[0]["c"]) if not df.empty else 0


def get_disqualified_pids(tournament_id: int) -> set:
    """Возвращает множество id дисквалифицированных участников турнира."""
    df = qdf("SELECT id FROM participants WHERE tournament_id=? AND COALESCE(disqualified,0)=1",
             (tournament_id,))
    return {int(r["id"]) for _, r in df.iterrows()} if not df.empty else set()


def set_participant_disqualified(participant_id: int, disqualified: bool):
    """Устанавливает или снимает дисквалификацию участника.
    При дисквалификации: авто-проставляет худший результат в квалификацию и во все heat_results."""
    exec_sql("UPDATE participants SET disqualified=? WHERE id=?",
             (1 if disqualified else 0, participant_id))
    if disqualified:
        # Получаем tournament_id и total_laps
        p_df = qdf("SELECT p.tournament_id, COALESCE(t.total_laps, 3) as total_laps "
                   "FROM participants p LEFT JOIN tournaments t ON t.id=p.tournament_id WHERE p.id=?",
                   (participant_id,))
        if not p_df.empty:
            tid = int(p_df.iloc[0]["tournament_id"])
            total_laps = int(p_df.iloc[0]["total_laps"])
            save_qual_result(tid, participant_id, 9999.0, 0.0, False, total_laps)
        # Обновляем все heat_results участника на худший результат
        exec_sql("""
            UPDATE heat_results SET time_seconds=9999, laps_completed=0, completed_all_laps=0,
                   place=4, points=0, projected_time=9999
            WHERE participant_id=? AND heat_id IN (
                SELECT h.id FROM heats h
                JOIN groups g ON g.id=h.group_id
                JOIN stages s ON s.id=g.stage_id
                WHERE s.tournament_id=(SELECT tournament_id FROM participants WHERE id=?)
            )
        """, (participant_id, participant_id))


def _apply_dsq_to_ranking(df: pd.DataFrame, tournament_id: int, pid_col: str = "pid") -> pd.DataFrame:
    """Перемещает дисквалифицированных участников в конец таблицы и пересчитывает места."""
    if df.empty:
        return df
    dsq_pids = get_disqualified_pids(tournament_id)
    if not dsq_pids:
        return df
    pid_name = pid_col if pid_col in df.columns else "participant_id"
    if pid_name not in df.columns:
        return df
    non_dsq = df[~df[pid_name].isin(dsq_pids)]
    dsq_rows = df[df[pid_name].isin(dsq_pids)]
    if dsq_rows.empty:
        return df
    combined = pd.concat([non_dsq, dsq_rows], ignore_index=True)
    if "rank" in combined.columns:
        combined["rank"] = range(1, len(combined) + 1)
    elif "place" in combined.columns:
        combined["place"] = range(1, len(combined) + 1)
    return combined.reset_index(drop=True)


# ============================================================
# Бизнес-логика: сетка и плей-офф
# ============================================================

def get_tournament(tournament_id: int) -> Optional[pd.Series]:
    df = qdf("SELECT * FROM tournaments WHERE id=?", (tournament_id,))
    return df.iloc[0] if not df.empty else None


def get_all_stages(tournament_id: int) -> pd.DataFrame:
    return qdf("SELECT * FROM stages WHERE tournament_id=? ORDER BY stage_idx", (tournament_id,))


def get_active_stage(tournament_id: int) -> Optional[pd.Series]:
    df = qdf("SELECT * FROM stages WHERE tournament_id=? AND status='active' ORDER BY stage_idx DESC LIMIT 1",
             (tournament_id,))
    return df.iloc[0] if not df.empty else None


def create_stage(tournament_id: int, stage_idx: int, sd: StageDef) -> int:
    exec_sql("""INSERT OR IGNORE INTO stages(tournament_id, stage_idx, code, group_size,
                group_count, qualifiers, heats_count, status)
                VALUES(?,?,?,?,?,?,?,'active')""",
             (tournament_id, stage_idx, sd.code, sd.group_size, sd.group_count,
              sd.qualifiers, sd.heats_count))
    stage_id = int(qdf("SELECT id FROM stages WHERE tournament_id=? AND stage_idx=?",
                        (tournament_id, stage_idx)).iloc[0]["id"])
    existing = int(qdf("SELECT COUNT(*) as c FROM groups WHERE stage_id=?", (stage_id,)).iloc[0]["c"])
    if existing == 0:
        exec_many("INSERT INTO groups(stage_id, group_no) VALUES(?,?)",
                  [(stage_id, gno) for gno in range(1, sd.group_count + 1)])
    return stage_id


def seed_groups_from_qual(tournament_id: int, stage_id: int, seeding_map: Dict[int, List[int]], advancing: int):
    """Посев из квалификации в первый этап плей-офф."""
    ranking = get_qual_ranking(tournament_id)
    if ranking.empty:
        return
    # Берём только прошедших
    ranking = ranking.head(advancing)

    groups_df = qdf("SELECT id, group_no FROM groups WHERE stage_id=?", (stage_id,))
    gid_by_no = {int(r["group_no"]): int(r["id"]) for _, r in groups_df.iterrows()}

    inserts = []
    for gno, seeds in seeding_map.items():
        for qual_rank in seeds:
            if qual_rank <= len(ranking):
                pid = int(ranking.iloc[qual_rank - 1]["pid"])
                inserts.append((gid_by_no[gno], pid))

    exec_many("INSERT OR IGNORE INTO group_members(group_id, participant_id) VALUES(?,?)", inserts)


def get_group_members(stage_id: int, group_no: int) -> pd.DataFrame:
    return qdf("""
        SELECT p.id as pid, p.name, p.start_number
        FROM groups g
        JOIN group_members gm ON gm.group_id=g.id
        JOIN participants p ON p.id=gm.participant_id
        WHERE g.stage_id=? AND g.group_no=?
        ORDER BY p.start_number
    """, (stage_id, int(group_no)))


def get_all_groups(stage_id: int) -> Dict[int, pd.DataFrame]:
    groups = qdf("SELECT group_no FROM groups WHERE stage_id=? ORDER BY group_no", (stage_id,))
    return {int(g["group_no"]): get_group_members(stage_id, int(g["group_no"])) for _, g in groups.iterrows()}


def save_heat(stage_id: int, group_no: int, heat_no: int, results: List[Dict],
              is_final: bool = False, track_no: int = 1, scoring: Optional[Dict] = None):
    """Сохранить результаты вылета. results = [{pid, time_seconds, laps_completed, completed_all_laps}]
    scoring: если передан, используется вместо FINAL_SCORING (например SIM_SCORING)."""
    group_id = int(qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?",
                        (stage_id, group_no)).iloc[0]["id"])
    exec_sql("INSERT OR IGNORE INTO heats(group_id, heat_no, track_no) VALUES(?,?,?)",
             (group_id, heat_no, track_no))
    heat_id = int(qdf("SELECT id FROM heats WHERE group_id=? AND heat_no=? AND track_no=?",
                       (group_id, heat_no, track_no)).iloc[0]["id"])

    # Ранжируем
    ranked = rank_results(results)

    tournament = qdf("SELECT t.total_laps, t.discipline FROM stages s JOIN tournaments t ON t.id=s.tournament_id WHERE s.id=?",
                      (stage_id,))
    total_laps = int(tournament.iloc[0]["total_laps"]) if not tournament.empty else 3
    disc = tournament.iloc[0]["discipline"] if not tournament.empty else "drone_individual"

    score_map = scoring if scoring is not None else (FINAL_SCORING if is_final else None)

    rows = []
    for r in ranked:
        if disc in ("sim_individual", "sim_team"):
            projected = None  # Нет расчётного времени для симулятора
        else:
            projected = calc_projected_time(r["time_seconds"], r["laps_completed"], total_laps) \
                if not r["completed_all_laps"] else r["time_seconds"]
        pts = score_map.get(r["place"], 0) if score_map else 0
        rows.append((heat_id, r["pid"], r["time_seconds"], r["laps_completed"],
                      int(r["completed_all_laps"]), projected, r["place"], pts))

    exec_sql("DELETE FROM heat_results WHERE heat_id=?", (heat_id,))
    exec_many("""INSERT INTO heat_results(heat_id, participant_id, time_seconds, laps_completed,
                 completed_all_laps, projected_time, place, points) VALUES(?,?,?,?,?,?,?,?)""", rows)


def get_heat_results(stage_id: int, group_no: int, heat_no: int, track_no: int = 1) -> List[Dict]:
    gid_df = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?", (stage_id, group_no))
    if gid_df.empty:
        return []
    group_id = int(gid_df.iloc[0]["id"])
    heat_df = qdf("SELECT id FROM heats WHERE group_id=? AND heat_no=? AND track_no=?",
                   (group_id, heat_no, track_no))
    if heat_df.empty:
        return []
    heat_id = int(heat_df.iloc[0]["id"])
    df = qdf("""SELECT hr.*, p.name, p.start_number FROM heat_results hr
                JOIN participants p ON p.id=hr.participant_id
                WHERE hr.heat_id=? ORDER BY hr.place""", (heat_id,))
    return df.to_dict("records") if not df.empty else []


def _tournament_id_from_stage(stage_id: int) -> Optional[int]:
    df = qdf("SELECT tournament_id FROM stages WHERE id=?", (stage_id,))
    return int(df.iloc[0]["tournament_id"]) if not df.empty else None


def compute_group_ranking(stage_id: int, group_no: int, discipline: str = "drone_individual",
                          scoring_mode: str = "none") -> pd.DataFrame:
    """Ранжирование в группе: для дронов — один вылет, для сима — агрегация + тайбрейк."""
    if discipline in ("sim_individual", "sim_team"):
        return resolve_sim_tiebreaker(stage_id, group_no, scoring_mode)
    results = get_heat_results(stage_id, group_no, 1)
    if not results:
        return pd.DataFrame()
    df = pd.DataFrame(results)
    tid = _tournament_id_from_stage(stage_id)
    return _apply_dsq_to_ranking(df, tid, "participant_id") if tid else df


def compute_final_standings(stage_id: int) -> pd.DataFrame:
    """Итоги финала: сумма очков за 3 основных вылета + бонус.
    Тайбрейкеры (вылеты 4+) используются только для разрешения ничьих."""
    group_id_df = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=1", (stage_id,))
    if group_id_df.empty:
        return pd.DataFrame()
    group_id = int(group_id_df.iloc[0]["id"])

    # Считаем очки только за основные 3 вылета
    df = qdf("""
        SELECT p.id as pid, p.name, p.start_number,
               COALESCE(SUM(hr.points), 0) as total_points,
               COALESCE(SUM(CASE WHEN hr.place=1 THEN 1 ELSE 0 END), 0) as wins,
               COUNT(hr.heat_id) as heats_played
        FROM group_members gm
        JOIN participants p ON p.id=gm.participant_id
        LEFT JOIN heats h ON h.group_id=gm.group_id AND h.heat_no <= 3
        LEFT JOIN heat_results hr ON hr.heat_id=h.id AND hr.participant_id=p.id
        WHERE gm.group_id=?
        GROUP BY p.id
    """, (group_id,))

    if df.empty:
        return df

    # Бонус +1 за 2+ побед
    df["bonus"] = (df["wins"] >= 2).astype(int)
    df["total"] = df["total_points"] + df["bonus"]

    # Узнаём максимальный номер тайбрейка
    max_heat_df = qdf("SELECT MAX(heat_no) as mx FROM heats WHERE group_id=?", (group_id,))
    max_heat = int(max_heat_df.iloc[0]["mx"]) if not max_heat_df.empty and max_heat_df.iloc[0]["mx"] is not None else 0

    # Строим ключ сортировки: total DESC, wins DESC, затем по тайбрейкерам (место ASC)
    df["tiebreak_key"] = 0  # чем меньше, тем лучше
    for tb_heat in range(4, max_heat + 1):
        tb_results = get_heat_results(stage_id, 1, tb_heat)
        if tb_results:
            tb_map = {r["participant_id"]: r["place"] for r in tb_results}
            # Для каждого тайбрейка добавляем место как вес (умножаем на убывающий коэффициент)
            col = f"tb_{tb_heat}"
            df[col] = df["pid"].map(lambda pid, m=tb_map: m.get(pid, 99))
            # Накапливаем: первый тайбрейк самый приоритетный
            df["tiebreak_key"] = df["tiebreak_key"] * 100 + df[col]

    df = df.sort_values(["total", "wins", "tiebreak_key"],
                        ascending=[False, False, True]).reset_index(drop=True)
    df["rank"] = range(1, len(df) + 1)

    # Определяем наличие ничьих (по total баллам, без учёта тайбрейков)
    df["has_tie"] = df.duplicated(subset=["total"], keep=False)

    tid = _tournament_id_from_stage(stage_id)
    return _apply_dsq_to_ranking(df, tid, "pid") if tid else df


def detect_final_ties(standings: pd.DataFrame) -> List[List[int]]:
    """Находит группы участников с одинаковыми очками, которые ещё не разрешены тайбрейком.
    Возвращает список групп pid'ов с ничьими."""
    if standings.empty:
        return []
    tied_groups = []
    for total_val, group in standings.groupby("total"):
        if len(group) > 1:
            # Проверяем, разрешена ли ничья тайбрейком
            if "tiebreak_key" in group.columns:
                tb_keys = group["tiebreak_key"].tolist()
                if len(set(tb_keys)) == len(tb_keys):
                    continue  # Все тайбрейки разные — ничья разрешена
            tied_groups.append(group["pid"].tolist())
    return tied_groups


def compute_sim_group_ranking(stage_id: int, group_no: int, scoring_mode: str = "sum_all") -> pd.DataFrame:
    """Ранжирование в группе для симулятора (2 трассы × 3 попытки).
    Сумма очков за все 6 вылетов. Макс 24 очка."""
    gid_df = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?", (stage_id, group_no))
    if gid_df.empty:
        return pd.DataFrame()
    group_id = int(gid_df.iloc[0]["id"])

    members = get_group_members(stage_id, group_no)
    if members.empty:
        return pd.DataFrame()

    # Сумма очков за все вылеты (исключая тайбрейки track_no=99)
    df = qdf("""
        SELECT hr.participant_id, p.name, p.start_number,
               COALESCE(SUM(hr.points), 0) as total_points,
               COUNT(hr.heat_id) as heats_played
        FROM heat_results hr
        JOIN heats h ON h.id=hr.heat_id
        JOIN participants p ON p.id=hr.participant_id
        WHERE h.group_id=? AND h.track_no < 99
        GROUP BY hr.participant_id
    """, (group_id,))
    if df.empty:
        return pd.DataFrame()
    df = df.sort_values("total_points", ascending=False).reset_index(drop=True)
    df["rank"] = range(1, len(df) + 1)
    tid = _tournament_id_from_stage(stage_id)
    return _apply_dsq_to_ranking(df, tid, "participant_id") if tid else df


def get_sim_track_bests(stage_id: int, group_no: int) -> Dict[int, Dict]:
    """Для каждого пилота в группе — лучшее время на Трассе 1 и Трассе 2.
    Возвращает {pid: {'t1': best_time_or_None, 't2': best_time_or_None}}."""
    gid_df = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?", (stage_id, group_no))
    if gid_df.empty:
        return {}
    group_id = int(gid_df.iloc[0]["id"])
    members = get_group_members(stage_id, group_no)
    result = {}
    for _, m in members.iterrows():
        pid = int(m["pid"])
        bests = {}
        for track in [1, 2]:
            bt = qdf("""
                SELECT MIN(hr.time_seconds) as best
                FROM heat_results hr
                JOIN heats h ON h.id=hr.heat_id
                WHERE h.group_id=? AND h.track_no=? AND hr.participant_id=?
                      AND hr.time_seconds > 0
            """, (group_id, track, pid))
            val = float(bt.iloc[0]["best"]) if not bt.empty and bt.iloc[0]["best"] is not None else None
            bests[f"t{track}"] = val
        result[pid] = bests
    return result


def detect_sim_group_ties(stage_id: int, group_no: int, scoring_mode: str, qualifiers: int = 2) -> List[List[int]]:
    """Обнаруживает критические ничьи в группе симулятора.
    Критическая ничья — когда пилоты на границе прохода имеют одинаковые очки
    (например, 2-е и 3-е место при qualifiers=2).
    Возвращает список групп pid'ов с критическими ничьими."""
    ranking = compute_sim_group_ranking(stage_id, group_no, scoring_mode)
    if ranking.empty or len(ranking) <= qualifiers:
        return []

    pid_col = "participant_id" if "participant_id" in ranking.columns else "pid"
    pts = ranking["total_points"].tolist()

    # Ищем ничью на границе: очки на позиции qualifiers-1 (последний проходящий)
    # совпадают с очками на позиции qualifiers (первый не проходящий)
    cutoff_pts = pts[qualifiers - 1]  # очки последнего проходящего
    tied_pids = []
    for i, p in enumerate(pts):
        if p == cutoff_pts:
            tied_pids.append(int(ranking.iloc[i][pid_col]))

    # Критическая ничья только если есть пилоты и по обе стороны границы
    has_above = any(i < qualifiers for i, p in enumerate(pts) if p == cutoff_pts)
    has_below = any(i >= qualifiers for i, p in enumerate(pts) if p == cutoff_pts)

    if has_above and has_below and len(tied_pids) > 1:
        return [tied_pids]
    return []


def resolve_sim_tiebreaker(stage_id: int, group_no: int, scoring_mode: str) -> pd.DataFrame:
    """Пересчитывает ранжирование группы с учётом тайбрейков (track_no=99)."""
    ranking = compute_sim_group_ranking(stage_id, group_no, scoring_mode)
    if ranking.empty:
        return ranking

    pid_col = "participant_id" if "participant_id" in ranking.columns else "pid"

    # Проверяем, есть ли тайбрейк-вылеты (track_no=99)
    tb_results = get_heat_results(stage_id, group_no, 1, track_no=99)
    if not tb_results:
        return ranking

    # Победитель тайбрейка (1-е место) получает +1 очко
    tb_rank = {r["participant_id"]: r["place"] for r in tb_results}
    for i, row in ranking.iterrows():
        pid = int(row[pid_col])
        if tb_rank.get(pid) == 1:
            ranking.at[i, "total_points"] = int(row["total_points"]) + 1

    # Пересортируем и при равных очках — по месту в тайбрейке
    ranking["tiebreak"] = ranking[pid_col].map(lambda x: tb_rank.get(int(x), 999))
    ranking = ranking.sort_values(["total_points", "tiebreak"], ascending=[False, True]).reset_index(drop=True)
    ranking["rank"] = range(1, len(ranking) + 1)
    ranking = ranking.drop(columns=["tiebreak"])
    tid = _tournament_id_from_stage(stage_id)
    return _apply_dsq_to_ranking(ranking, tid, pid_col) if tid else ranking


def compute_sim_final_standings(stage_id: int, scoring_mode: str) -> pd.DataFrame:
    """Итоги финала симулятора: те же правила, что и групповой этап, но без бонусов."""
    return resolve_sim_tiebreaker(stage_id, 1, scoring_mode)


def compute_overall_standings(tournament_id: int) -> pd.DataFrame:
    """Вычисляет общую итоговую таблицу турнира: каждый участник получает своё место.

    Логика распределения мест:
    1. Финалисты: места 1-4 из итогов финала
    2. Проигравшие полуфинала: места 5-8 (ранжированы внутри)
    3. Проигравшие четвертьфинала: места 9-16
    4. Проигравшие 1/8: места 17-32
    5. Не прошедшие квалификацию: следующие места
    """
    tourn = get_tournament(tournament_id)
    disc = str(tourn["discipline"])
    sm = str(tourn.get("scoring_mode", "none"))
    is_sim_ov = disc in ("sim_individual", "sim_team")

    bracket = get_bracket_for_tournament(tournament_id)
    all_stages = get_all_stages(tournament_id)

    overall = []  # список {place, pid, name, stage_eliminated, detail}
    placed_pids = set()
    current_place = 1

    if bracket and not all_stages.empty:
        # Проходим этапы с конца (финал → ... → первый этап)
        for sidx in range(len(bracket) - 1, -1, -1):
            sd = bracket[sidx]
            srow = all_stages[all_stages["stage_idx"] == sidx]
            if srow.empty:
                continue
            stage_id = int(srow.iloc[0]["id"])
            sname = sd.display_name.get("ru", sd.code)

            if sd.code == "F":
                # Финалисты — из итогов финала
                if is_sim_ov:
                    fin = compute_sim_final_standings(stage_id, sm)
                    pid_col = "participant_id" if "participant_id" in fin.columns else "pid"
                    if not fin.empty:
                        for _, row in fin.iterrows():
                            pid = int(row[pid_col])
                            detail = f"{int(row['total_points'])} оч."
                            overall.append({
                                "place": int(row["rank"]),
                                "pid": pid, "name": row["name"],
                                "stage": sname, "detail": detail,
                            })
                            placed_pids.add(pid)
                        current_place = len(fin) + 1
                else:
                    fin = compute_final_standings(stage_id)
                    if not fin.empty:
                        for _, row in fin.iterrows():
                            pid = int(row["pid"])
                            detail = f"{int(row['total'])} оч. ({int(row['wins'])} поб.)"
                            overall.append({
                                "place": int(row["rank"]),
                                "pid": pid, "name": row["name"],
                                "stage": sname, "detail": detail,
                            })
                            placed_pids.add(pid)
                        current_place = len(fin) + 1
            else:
                # Не финальный этап: определяем кто выбыл (3-е и 4-е место в группах)
                all_groups_ov = get_all_groups(stage_id)
                eliminated = []

                for gno in sorted(all_groups_ov.keys()):
                    if is_sim_ov:
                        ranking = compute_sim_group_ranking(stage_id, gno, sm)
                        pid_col = "participant_id" if "participant_id" in ranking.columns else "pid"
                        if not ranking.empty:
                            for _, row in ranking.iterrows():
                                pid = int(row[pid_col])
                                if pid not in placed_pids and int(row["rank"]) > sd.qualifiers:
                                    eliminated.append({
                                        "pid": pid, "name": row["name"],
                                        "sort_key": int(row["total_points"]),
                                        "detail": f"{int(row['total_points'])} оч.",
                                        "rank_in_group": int(row["rank"]),
                                    })
                    else:
                        results = get_heat_results(stage_id, gno, 1)
                        if not results:
                            # Fallback
                            gid_fb = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?", (stage_id, gno))
                            if not gid_fb.empty:
                                gid_val = int(gid_fb.iloc[0]["id"])
                                fb_heat = qdf("SELECT id FROM heats WHERE group_id=? ORDER BY heat_no LIMIT 1", (gid_val,))
                                if not fb_heat.empty:
                                    fb_hid = int(fb_heat.iloc[0]["id"])
                                    fb_df = qdf("""SELECT hr.*, p.name, p.start_number FROM heat_results hr
                                                   JOIN participants p ON p.id=hr.participant_id
                                                   WHERE hr.heat_id=? ORDER BY hr.place""", (fb_hid,))
                                    if not fb_df.empty:
                                        results = fb_df.to_dict("records")
                        if results:
                            for r in results:
                                pid = int(r["participant_id"])
                                if pid not in placed_pids and r["place"] > sd.qualifiers:
                                    t = r.get("projected_time") or r.get("time_seconds")
                                    eliminated.append({
                                        "pid": pid, "name": r["name"],
                                        "sort_key": t if t else 9999,
                                        "detail": format_time(t),
                                        "rank_in_group": r["place"],
                                    })

                # Сортируем выбывших: для сима — по очкам DESC, для дронов — по времени ASC
                if is_sim_ov:
                    eliminated.sort(key=lambda x: -x["sort_key"])
                else:
                    eliminated.sort(key=lambda x: x["sort_key"])

                # Дисквалифицированные всегда в конце списка выбывших
                dsq_pids = get_disqualified_pids(tournament_id)
                if dsq_pids:
                    non_dsq = [e for e in eliminated if e["pid"] not in dsq_pids]
                    dsq_list = [e for e in eliminated if e["pid"] in dsq_pids]
                    eliminated = non_dsq + dsq_list

                for e in eliminated:
                    overall.append({
                        "place": current_place,
                        "pid": e["pid"], "name": e["name"],
                        "stage": sname, "detail": e["detail"],
                    })
                    placed_pids.add(e["pid"])
                    current_place += 1

    # Не прошедшие квалификацию
    qual_ranking = get_qual_ranking(tournament_id)
    if not qual_ranking.empty:
        advancing = compute_bracket_size(len(qual_ranking))
        for _, row in qual_ranking.iterrows():
            pid = int(row["pid"])
            if pid not in placed_pids:
                t = row.get("projected_time") or row.get("time_seconds")
                overall.append({
                    "place": current_place,
                    "pid": pid, "name": row["name"],
                    "stage": "Квалификация",
                    "detail": format_time(t) if t and pd.notna(t) else "—",
                })
                placed_pids.add(pid)
                current_place += 1

    # Финальная пересортировка: дисквалифицированные — в конец каждого блока этапа
    if overall:
        dsq_pids = get_disqualified_pids(tournament_id)
        if dsq_pids:
            result = []
            i = 0
            while i < len(overall):
                stage = overall[i]["stage"]
                block = []
                while i < len(overall) and overall[i]["stage"] == stage:
                    block.append(overall[i].copy())
                    i += 1
                non_dsq = [e for e in block if e["pid"] not in dsq_pids]
                dsq_block = [e for e in block if e["pid"] in dsq_pids]
                block = non_dsq + dsq_block
                result.extend(block)
            for idx, r in enumerate(result):
                r["place"] = idx + 1
            overall = result
        return pd.DataFrame(overall)
    return pd.DataFrame()


def check_stage_results_complete(stage_id: int, stage_def: StageDef, disc: str = "drone_individual",
                                 scoring_mode: str = "none") -> Tuple[bool, str]:
    """Проверяет, все ли результаты заполнены для текущего этапа.
    Возвращает (ok, message)."""
    all_groups = get_all_groups(stage_id)
    if not all_groups:
        return False, "Нет групп в этом этапе"

    missing = []
    for gno, members in all_groups.items():
        if members.empty:
            missing.append(f"Группа {gno}: нет участников")
            continue

        if disc in ("sim_individual", "sim_team"):
            # Для симулятора: 2 трассы × 3 попытки = 6 вылетов
            for track in [1, 2]:
                for attempt in [1, 2, 3]:
                    results = get_heat_results(stage_id, gno, attempt, track)
                    if not results:
                        missing.append(f"Группа {gno}, Трасса {track}, Попытка {attempt}: нет результатов")
            # Проверяем неразрешённые ничьи
            if not missing:  # только если все результаты заполнены
                tied = detect_sim_group_ties(stage_id, gno, scoring_mode, stage_def.qualifiers)
                # Проверяем, разрешена ли ничья тайбрейком
                if tied:
                    tb = get_heat_results(stage_id, gno, 1, track_no=99)
                    if not tb:
                        tied_ranking = compute_sim_group_ranking(stage_id, gno, scoring_mode)
                        pid_col = "participant_id" if "participant_id" in tied_ranking.columns else "pid"
                        for tg in tied:
                            names = tied_ranking[tied_ranking[pid_col].isin(tg)]["name"].tolist()
                            missing.append(f"Группа {gno}: ничья между {', '.join(names)} — нужен доп. вылет")
        else:
            # Для дронов: heats_count вылетов (1 для плей-офф, 3 для финала)
            heats_needed = stage_def.heats_count
            for h in range(1, heats_needed + 1):
                results = get_heat_results(stage_id, gno, h)
                if not results:
                    if heats_needed > 1:
                        missing.append(f"Группа {gno}, вылет {h}: нет результатов")
                    else:
                        missing.append(f"Группа {gno}: нет результатов")
    if missing:
        return False, "Не все результаты заполнены:\n" + "\n".join(missing)
    return True, ""


def advance_to_next_stage(tournament_id: int, bracket: List[StageDef]):
    """Переход к следующему этапу плей-офф."""
    stages_df = get_all_stages(tournament_id)
    active = stages_df[stages_df["status"] == "active"]
    if active.empty:
        return
    cur = active.iloc[0]
    cur_idx = int(cur["stage_idx"])
    next_idx = cur_idx + 1
    if next_idx >= len(bracket):
        return

    # Получаем дисциплину и scoring_mode
    tourn_info = qdf("SELECT discipline, scoring_mode FROM tournaments WHERE id=?", (tournament_id,))
    disc = str(tourn_info.iloc[0]["discipline"]) if not tourn_info.empty else "drone_individual"
    sm = str(tourn_info.iloc[0]["scoring_mode"]) if not tourn_info.empty else "none"

    # Валидация: проверить, что все результаты текущего этапа заполнены
    cur_sd = bracket[cur_idx]
    ok, msg = check_stage_results_complete(int(cur["id"]), cur_sd, disc, sm)
    if not ok:
        raise ValueError(msg)

    next_sd = bracket[next_idx]
    next_stage_id = create_stage(tournament_id, next_idx, next_sd)

    if next_sd.progress_map:
        groups_df = qdf("SELECT id, group_no FROM groups WHERE stage_id=?", (next_stage_id,))
        gid_by_no = {int(r["group_no"]): int(r["id"]) for _, r in groups_df.iterrows()}

        rows = []
        for target_gno, refs in next_sd.progress_map.items():
            for (place, src_gno) in refs:
                ranking = compute_group_ranking(int(cur["id"]), src_gno, disc, sm)
                if not ranking.empty and len(ranking) >= place:
                    # pid column name can differ between drone and sim ranking DataFrames
                    pid_col = "participant_id" if "participant_id" in ranking.columns else "pid"
                    pid = int(ranking.iloc[place - 1][pid_col])
                    rows.append((gid_by_no[target_gno], pid))

        exec_many("INSERT OR IGNORE INTO group_members(group_id, participant_id) VALUES(?,?)", rows)

    exec_sql("UPDATE stages SET status='done' WHERE id=?", (int(cur["id"]),))


def rollback_to_previous_stage(tournament_id: int, bracket: List[StageDef]):
    """Откат на предыдущий этап: удаляет текущий активный этап и реактивирует предыдущий."""
    stages_df = get_all_stages(tournament_id)

    # Если турнир завершён — снимаем статус finished и реактивируем финал
    tourn = get_tournament(tournament_id)
    if str(tourn["status"]) == "finished":
        # Находим последний этап (финал)
        last = stages_df[stages_df["status"] == "done"].sort_values("stage_idx", ascending=False)
        if not last.empty:
            last_id = int(last.iloc[0]["id"])
            exec_sql("UPDATE stages SET status='active' WHERE id=?", (last_id,))
            exec_sql("UPDATE tournaments SET status='bracket' WHERE id=?", (tournament_id,))
        return

    active = stages_df[stages_df["status"] == "active"]
    if active.empty:
        return
    cur = active.iloc[0]
    cur_idx = int(cur["stage_idx"])
    cur_stage_id = int(cur["id"])

    if cur_idx == 0:
        # Первый этап — откатываемся в квалификацию
        # Удаляем все данные этапа
        groups = qdf("SELECT id FROM groups WHERE stage_id=?", (cur_stage_id,))
        for _, g in groups.iterrows():
            gid = int(g["id"])
            heats = qdf("SELECT id FROM heats WHERE group_id=?", (gid,))
            for _, h in heats.iterrows():
                exec_sql("DELETE FROM heat_results WHERE heat_id=?", (int(h["id"]),))
            exec_sql("DELETE FROM heats WHERE group_id=?", (gid,))
            exec_sql("DELETE FROM group_members WHERE group_id=?", (gid,))
        exec_sql("DELETE FROM groups WHERE stage_id=?", (cur_stage_id,))
        exec_sql("DELETE FROM stages WHERE id=?", (cur_stage_id,))
        exec_sql("UPDATE tournaments SET status='qualification' WHERE id=?", (tournament_id,))
    else:
        # Удаляем текущий этап и реактивируем предыдущий
        groups = qdf("SELECT id FROM groups WHERE stage_id=?", (cur_stage_id,))
        for _, g in groups.iterrows():
            gid = int(g["id"])
            heats = qdf("SELECT id FROM heats WHERE group_id=?", (gid,))
            for _, h in heats.iterrows():
                exec_sql("DELETE FROM heat_results WHERE heat_id=?", (int(h["id"]),))
            exec_sql("DELETE FROM heats WHERE group_id=?", (gid,))
            exec_sql("DELETE FROM group_members WHERE group_id=?", (gid,))
        exec_sql("DELETE FROM groups WHERE stage_id=?", (cur_stage_id,))
        exec_sql("DELETE FROM stages WHERE id=?", (cur_stage_id,))

        # Реактивируем предыдущий
        prev = stages_df[stages_df["stage_idx"] == cur_idx - 1]
        if not prev.empty:
            exec_sql("UPDATE stages SET status='active' WHERE id=?", (int(prev.iloc[0]["id"]),))


def start_bracket(tournament_id: int):
    """Завершить квалификацию и создать первый этап плей-офф."""
    ranking = get_qual_ranking(tournament_id)
    n = len(ranking)
    advancing = compute_bracket_size(n)
    bracket = generate_bracket(advancing)

    # Сохраняем bracket info
    exec_sql("UPDATE tournaments SET status='bracket' WHERE id=?", (tournament_id,))

    # Создаём первый этап
    first_sd = bracket[0]
    stage_id = create_stage(tournament_id, 0, first_sd)

    # Посев
    if first_sd.seeding_map:
        seed_groups_from_qual(tournament_id, stage_id, first_sd.seeding_map, advancing)


def get_bracket_for_tournament(tournament_id: int) -> List[StageDef]:
    """Определяет сетку по количеству прошедших квалификацию."""
    ranking = get_qual_ranking(tournament_id)
    n = len(ranking)
    if n == 0:
        # Проверяем если уже есть этапы — восстанавливаем по кол-ву участников первого этапа
        stages = get_all_stages(tournament_id)
        if not stages.empty:
            first = stages.iloc[0]
            total = int(first["group_size"]) * int(first["group_count"])
            return generate_bracket(total)
        return []
    advancing = compute_bracket_size(n)
    return generate_bracket(advancing)


# ============================================================
# UI-хелперы
# ============================================================

def style_qual_table(df: pd.DataFrame, cutoff: int):
    """Подсветка: зелёный для проходящих, красный для отсечённых."""
    def highlight(row):
        rank = row.name + 1  # 0-based index
        if rank <= cutoff:
            return ["background-color: #1a472a; color: #90EE90"] * len(row)
        else:
            return ["background-color: #4a1a1a; color: #FFB6B6"] * len(row)
    return df.style.apply(highlight, axis=1)


def style_standings_table(df: pd.DataFrame, qualifiers: int):
    """Подсветка для плей-офф таблиц."""
    def highlight_row(row):
        rank = row["М"]
        if qualifiers > 0:
            if rank <= qualifiers:
                return ["background-color: #1a472a; color: #90EE90"] * len(row)
            else:
                return ["background-color: #4a1a1a; color: #FFB6B6"] * len(row)
        return [""] * len(row)
    return df.style.apply(highlight_row, axis=1)


def style_final_podium(df: pd.DataFrame):
    """Подсветка финальной таблицы: золото, серебро, бронза."""
    MEDAL_COLORS = {
        1: "background-color: #5C4B00; color: #FFD700",   # Золото
        2: "background-color: #3A3A3A; color: #C0C0C0",   # Серебро
        3: "background-color: #3D2B1F; color: #CD7F32",   # Бронза
    }

    def highlight_row(row):
        rank = row["М"]
        if rank in MEDAL_COLORS:
            return [MEDAL_COLORS[rank]] * len(row)
        return [""] * len(row)
    return df.style.apply(highlight_row, axis=1)


def download_csv_button(df: pd.DataFrame, label: str, filename: str):
    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(label, data=csv, file_name=filename, mime="text/csv")


def export_tournament_excel(tournament_id: int) -> bytes:
    """Генерирует полный Excel-отчёт по турниру (все этапы)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    tourn = get_tournament(tournament_id)
    disc = str(tourn["discipline"])
    is_sim_export = disc in ("sim_individual", "sim_team")
    is_team_export = disc == "sim_team"
    scoring_mode_exp = str(tourn.get("scoring_mode", "none"))
    time_limit_exp = float(tourn["time_limit_seconds"])
    total_laps_exp = int(tourn["total_laps"])

    DISCIPLINES_RU = {
        "drone_individual": "Дроны: Личный зачёт",
        "sim_individual": "Симулятор: Личный зачёт",
        "sim_team": "Симулятор: Командный зачёт",
    }

    wb = Workbook()
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill("solid", fgColor="D5F5E3")
    red_fill = PatternFill("solid", fgColor="FADBD8")
    gold_fill = PatternFill("solid", fgColor="F9E79F")
    silver_fill = PatternFill("solid", fgColor="D5DBDB")
    bronze_fill = PatternFill("solid", fgColor="E8DAEF")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill("solid", fgColor="D6EAF8")

    def style_header(ws, row_num, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_row(ws, row_num, max_col, fill=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.alignment = cell_align
            cell.border = thin_border
            if fill:
                cell.fill = fill

    def auto_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def write_df(ws, df, start_row=1, advancing=None, medal_col=False):
        """Записывает DataFrame на лист с форматированием."""
        headers = list(df.columns)
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=c_idx, value=h)
        style_header(ws, start_row, len(headers))

        for r_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
            fill = None
            if advancing is not None and "М" in df.columns:
                place = row.get("М", r_idx - start_row)
                if isinstance(place, (int, float)) and place <= advancing:
                    fill = green_fill
                elif isinstance(place, (int, float)):
                    fill = red_fill
            if medal_col and "М" in df.columns:
                place = row.get("М", 999)
                if place == 1:
                    fill = gold_fill
                elif place == 2:
                    fill = silver_fill
                elif place == 3:
                    fill = bronze_fill
            for c_idx, h in enumerate(headers, 1):
                ws.cell(row=r_idx, column=c_idx, value=row[h])
            style_data_row(ws, r_idx, len(headers), fill)
        return start_row + len(df) + 1

    # ===== Лист 1: Информация о турнире =====
    ws_info = wb.active
    ws_info.title = "Турнир"
    info_data = [
        ("Турнир", str(tourn["name"])),
        ("Дисциплина", DISCIPLINES_RU.get(disc, disc)),
        ("Статус", str(tourn["status"])),
        ("Лимит времени (сек)", time_limit_exp),
        ("Кругов", total_laps_exp),
        ("Дата создания", str(tourn.get("created_at", "—"))),
    ]
    if is_sim_export:
        info_data.append(("Подсчёт очков", "Сумма за 6 вылетов (2 трассы × 3 попытки)"))
    if is_team_export:
        info_data.append(("Формат", "Командный (2 пилота, время суммируется)"))
    for r, (label, val) in enumerate(info_data, 1):
        ws_info.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_info.cell(row=r, column=2, value=val)
    auto_width(ws_info)

    # ===== Лист 2: Участники =====
    ws_part = wb.create_sheet("Участники")
    participants_raw = qdf("""SELECT id, start_number, name
                              FROM participants WHERE tournament_id=?
                              ORDER BY COALESCE(start_number, 9999), name""", (tournament_id,))

    if not participants_raw.empty:
        if is_team_export:
            tp_df = qdf("SELECT participant_id, pilot1_name, pilot2_name FROM team_pilots WHERE participant_id IN ({})".format(
                ",".join(str(int(x)) for x in participants_raw["id"].tolist())))
            tp_map = {}
            for _, tpr in tp_df.iterrows():
                tp_map[int(tpr["participant_id"])] = (tpr["pilot1_name"], tpr["pilot2_name"])
            rows = []
            for _, r in participants_raw.iterrows():
                pid = int(r["id"])
                pilots = tp_map.get(pid, ("", ""))
                rows.append({
                    "№": int(r["start_number"]) if pd.notna(r["start_number"]) else "",
                    "Команда": r["name"],
                    "Пилот 1": pilots[0],
                    "Пилот 2": pilots[1],
                })
            df_p = pd.DataFrame(rows)
        else:
            rows = []
            for _, r in participants_raw.iterrows():
                rows.append({
                    "№": int(r["start_number"]) if pd.notna(r["start_number"]) else "",
                    "Пилот": r["name"],
                })
            df_p = pd.DataFrame(rows)
        write_df(ws_part, df_p)
    auto_width(ws_part)

    # ===== Лист 3: Квалификация =====
    ws_qual = wb.create_sheet("Квалификация")
    ranking = get_qual_ranking(tournament_id)
    if not ranking.empty:
        advancing = compute_bracket_size(len(ranking))
        entity = "Команда" if is_team_export else "Пилот"

        if is_sim_export:
            qdf_data = ranking[["place", "name", "start_number", "time_seconds", "laps_completed"]].copy()
            qdf_data.columns = ["М", entity, "№", "Время (сек)", "Круги"]
        else:
            qdf_data = ranking[["place", "name", "start_number", "time_seconds",
                                "laps_completed", "completed_all_laps", "projected_time"]].copy()
            qdf_data.columns = ["М", entity, "№", "Время (сек)", "Круги", "Все круги", "Расчётное время"]

        # Заполняем расчётное форматированным
        if "Расчётное время" in qdf_data.columns:
            qdf_data["Расчётное время"] = qdf_data["Расчётное время"].apply(
                lambda x: format_time(x) if pd.notna(x) else "—")

        ws_qual.cell(row=1, column=1, value=f"Проходят: {advancing} из {len(ranking)}").font = Font(bold=True, size=11)
        write_df(ws_qual, qdf_data, start_row=3, advancing=advancing)
    auto_width(ws_qual)

    # ===== Листы 4+: Этапы плей-офф / группового этапа =====
    bracket_exp = get_bracket_for_tournament(tournament_id)
    all_stages_exp = get_all_stages(tournament_id)

    if bracket_exp and not all_stages_exp.empty:
        for sidx, sd in enumerate(bracket_exp):
            srow = all_stages_exp[all_stages_exp["stage_idx"] == sidx]
            if srow.empty:
                continue
            stage_id_exp = int(srow.iloc[0]["id"])
            sname = sd.display_name.get("ru", sd.code)
            is_final_stage = sd.code == "F"

            # Безопасное имя листа (макс 31 символ)
            # Excel не допускает символы / \ ? * [ ] : в названии листа
            sheet_name = sname.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")").replace(":", "-")[:31]
            ws_stage = wb.create_sheet(sheet_name)
            current_row = 1

            if is_final_stage:
                # === ФИНАЛ ===
                ws_stage.cell(row=current_row, column=1, value=f"ФИНАЛ — {sname}").font = Font(bold=True, size=13)
                current_row += 2

                if is_sim_export:
                    # Sim final: show per-track/attempt results + доп. вылеты + final standings
                    for tr in [1, 2]:
                        for att in [1, 2, 3]:
                            results = get_heat_results(stage_id_exp, 1, att, tr)
                            if results:
                                ent = "Команда" if is_team_export else "Пилот"
                                ws_stage.cell(row=current_row, column=1,
                                              value=f"Трасса {tr}, Попытка {att}").font = subheader_font
                                ws_stage.cell(row=current_row, column=1).fill = subheader_fill
                                current_row += 1
                                tdata = [{"М": r["place"], ent: r["name"],
                                          "Время": format_time(r.get("time_seconds")),
                                          "Круги": r.get("laps_completed", "—"),
                                          "Очки": int(r.get("points", 0))} for r in results]
                                df_heat = pd.DataFrame(tdata)
                                current_row = write_df(ws_stage, df_heat, current_row)
                                current_row += 1

                    # Доп. вылеты (тайбрейкеры) для симулятора
                    gid_sim_fin = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=1", (stage_id_exp,))
                    if not gid_sim_fin.empty:
                        max_heat_sim = qdf("SELECT MAX(heat_no) as mx FROM heats WHERE group_id=? AND track_no=1",
                                           (int(gid_sim_fin.iloc[0]["id"]),))
                        max_hn_sim = int(max_heat_sim.iloc[0]["mx"]) if not max_heat_sim.empty and max_heat_sim.iloc[0]["mx"] is not None else 0
                        for tb_att in range(4, max_hn_sim + 1):
                            tb_res = get_heat_results(stage_id_exp, 1, tb_att, track_no=1)
                            if tb_res:
                                ent = "Команда" if is_team_export else "Пилот"
                                ws_stage.cell(row=current_row, column=1,
                                              value=f"Доп. вылет #{tb_att - 3}").font = subheader_font
                                ws_stage.cell(row=current_row, column=1).fill = subheader_fill
                                current_row += 1
                                tdata = [{"М": r["place"], ent: r["name"],
                                          "Время": format_time(r.get("time_seconds")),
                                          "Круги": r.get("laps_completed", "—"),
                                          "Очки": int(r.get("points", 0))} for r in tb_res]
                                df_heat = pd.DataFrame(tdata)
                                current_row = write_df(ws_stage, df_heat, current_row)
                                current_row += 1

                    # Final standings
                    ws_stage.cell(row=current_row, column=1, value="ИТОГО ФИНАЛА").font = Font(bold=True, size=12)
                    current_row += 1
                    sim_fin = compute_sim_final_standings(stage_id_exp, scoring_mode_exp)
                    if not sim_fin.empty:
                        track_bests = get_sim_track_bests(stage_id_exp, 1)
                        pid_col = "participant_id" if "participant_id" in sim_fin.columns else "pid"
                        ent = "Команда" if is_team_export else "Пилот"
                        fin_rows = []
                        for _, row in sim_fin.iterrows():
                            pid = int(row[pid_col])
                            tb = track_bests.get(pid, {})
                            t1 = format_time(tb.get("t1")) if tb.get("t1") else "—"
                            t2 = format_time(tb.get("t2")) if tb.get("t2") else "—"
                            fin_rows.append({
                                "М": int(row["rank"]), ent: row["name"],
                                "Лучш. Тр.1": t1, "Лучш. Тр.2": t2,
                                "Очки": int(row["total_points"]),
                            })
                        df_fin = pd.DataFrame(fin_rows)
                        current_row = write_df(ws_stage, df_fin, current_row, medal_col=True)
                else:
                    # Drone final: 3 heats + доп. вылеты + standings
                    for heat_no in range(1, 4):
                        results = get_heat_results(stage_id_exp, 1, heat_no)
                        if results:
                            ws_stage.cell(row=current_row, column=1,
                                          value=f"Вылет {heat_no}").font = subheader_font
                            ws_stage.cell(row=current_row, column=1).fill = subheader_fill
                            current_row += 1
                            tdata = [{"М": r["place"], "Пилот": r["name"],
                                      "Время": format_time(r.get("time_seconds")),
                                      "Круги": r.get("laps_completed", "—"),
                                      "Все": "Да" if r.get("completed_all_laps") else "",
                                      "Расчётное": format_time(r.get("projected_time")),
                                      "Очки": int(r.get("points", 0))} for r in results]
                            df_heat = pd.DataFrame(tdata)
                            current_row = write_df(ws_stage, df_heat, current_row)
                            current_row += 1

                    # Доп. вылеты (тайбрейкеры) для дронов
                    gid_fin = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=1", (stage_id_exp,))
                    if not gid_fin.empty:
                        max_heat_fin = qdf("SELECT MAX(heat_no) as mx FROM heats WHERE group_id=?",
                                           (int(gid_fin.iloc[0]["id"]),))
                        max_hn = int(max_heat_fin.iloc[0]["mx"]) if not max_heat_fin.empty and max_heat_fin.iloc[0]["mx"] is not None else 0
                        for tb_no in range(4, max_hn + 1):
                            tb_res = get_heat_results(stage_id_exp, 1, tb_no)
                            if tb_res:
                                ws_stage.cell(row=current_row, column=1,
                                              value=f"Доп. вылет #{tb_no - 3}").font = subheader_font
                                ws_stage.cell(row=current_row, column=1).fill = subheader_fill
                                current_row += 1
                                tdata = [{"М": r["place"], "Пилот": r["name"],
                                          "Время": format_time(r.get("time_seconds")),
                                          "Круги": r.get("laps_completed", "—"),
                                          "Все": "Да" if r.get("completed_all_laps") else "",
                                          "Расчётное": format_time(r.get("projected_time")),
                                          "Очки": int(r.get("points", 0))} for r in tb_res]
                                df_heat = pd.DataFrame(tdata)
                                current_row = write_df(ws_stage, df_heat, current_row)
                                current_row += 1

                    # Drone final standings
                    ws_stage.cell(row=current_row, column=1, value="ИТОГО ФИНАЛА").font = Font(bold=True, size=12)
                    current_row += 1
                    fin_standings = compute_final_standings(stage_id_exp)
                    if not fin_standings.empty:
                        fin_rows = []
                        for _, row in fin_standings.iterrows():
                            fin_rows.append({
                                "М": int(row["rank"]),
                                "Пилот": row["name"],
                                "Очки": int(row["total_points"]),
                                "Побед": int(row["wins"]),
                                "Бонус": "+1" if int(row.get("bonus", 0)) > 0 else "",
                                "Итого": int(row["total"]),
                            })
                        df_fin = pd.DataFrame(fin_rows)
                        current_row = write_df(ws_stage, df_fin, current_row, medal_col=True)
            else:
                # === ГРУППОВОЙ ЭТАП / ПЛЕЙ-ОФФ ===
                ws_stage.cell(row=current_row, column=1, value=sname).font = Font(bold=True, size=13)
                current_row += 2

                all_groups_exp = get_all_groups(stage_id_exp)
                for gno, members in sorted(all_groups_exp.items()):
                    ws_stage.cell(row=current_row, column=1,
                                  value=f"Группа {gno}").font = subheader_font
                    ws_stage.cell(row=current_row, column=1).fill = subheader_fill
                    current_row += 1

                    ent = "Команда" if is_team_export else "Пилот"

                    if is_sim_export:
                        # Per-track/attempt results
                        for tr in [1, 2]:
                            for att in [1, 2, 3]:
                                results = get_heat_results(stage_id_exp, gno, att, tr)
                                if results:
                                    ws_stage.cell(row=current_row, column=1,
                                                  value=f"  Тр.{tr} Поп.{att}").font = Font(italic=True)
                                    current_row += 1
                                    tdata = [{"М": r["place"], ent: r["name"],
                                              "Время": format_time(r.get("time_seconds")),
                                              "Очки": int(r.get("points", 0))} for r in results]
                                    df_heat = pd.DataFrame(tdata)
                                    current_row = write_df(ws_stage, df_heat, current_row)
                                    current_row += 1

                        # Доп. вылет (тайбрейк) для плей-офф симулятора
                        tb_po = get_heat_results(stage_id_exp, gno, 1, track_no=99)
                        if tb_po:
                            ws_stage.cell(row=current_row, column=1,
                                          value=f"  Доп. вылет (тайбрейк)").font = Font(italic=True)
                            current_row += 1
                            tdata = [{"М": r["place"], ent: r["name"],
                                      "Время": format_time(r.get("time_seconds")),
                                      "Очки": int(r.get("points", 0))} for r in tb_po]
                            df_heat = pd.DataFrame(tdata)
                            current_row = write_df(ws_stage, df_heat, current_row)
                            current_row += 1

                        # Group summary
                        sim_rank = compute_sim_group_ranking(stage_id_exp, gno, scoring_mode_exp)
                        if not sim_rank.empty:
                            track_bests = get_sim_track_bests(stage_id_exp, gno)
                            pid_col = "participant_id" if "participant_id" in sim_rank.columns else "pid"
                            ws_stage.cell(row=current_row, column=1,
                                          value=f"  Сводка группы {gno}").font = Font(bold=True, italic=True)
                            current_row += 1
                            sum_rows = []
                            for _, sr in sim_rank.iterrows():
                                pid = int(sr[pid_col])
                                tb = track_bests.get(pid, {})
                                t1 = format_time(tb.get("t1")) if tb.get("t1") else "—"
                                t2 = format_time(tb.get("t2")) if tb.get("t2") else "—"
                                sum_rows.append({
                                    "М": int(sr["rank"]), ent: sr["name"],
                                    "Лучш. Тр.1": t1, "Лучш. Тр.2": t2,
                                    "Очки": int(sr["total_points"]),
                                })
                            df_sum = pd.DataFrame(sum_rows)
                            current_row = write_df(ws_stage, df_sum, current_row, advancing=sd.qualifiers)
                    else:
                        # Drone: single heat per group — пробуем прямой запрос
                        results = get_heat_results(stage_id_exp, gno, 1)
                        if not results:
                            # Fallback: может быть track_no другой — ищем любой heat
                            gid_fb = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?", (stage_id_exp, gno))
                            if not gid_fb.empty:
                                gid_val = int(gid_fb.iloc[0]["id"])
                                fb_heat = qdf("SELECT id, heat_no FROM heats WHERE group_id=? ORDER BY heat_no LIMIT 1", (gid_val,))
                                if not fb_heat.empty:
                                    fb_hid = int(fb_heat.iloc[0]["id"])
                                    fb_df = qdf("""SELECT hr.*, p.name, p.start_number FROM heat_results hr
                                                   JOIN participants p ON p.id=hr.participant_id
                                                   WHERE hr.heat_id=? ORDER BY hr.place""", (fb_hid,))
                                    if not fb_df.empty:
                                        results = fb_df.to_dict("records")
                        if results:
                            tdata = [{"М": r["place"], ent: r["name"],
                                      "Время": format_time(r.get("time_seconds")),
                                      "Круги": r.get("laps_completed", "—"),
                                      "Все": "Да" if r.get("completed_all_laps") else "",
                                      "Расчётное": format_time(r.get("projected_time"))} for r in results]
                            df_heat = pd.DataFrame(tdata)
                            current_row = write_df(ws_stage, df_heat, current_row, advancing=sd.qualifiers)
                        else:
                            # Последний fallback — хотя бы список участников группы
                            if not members.empty:
                                tdata = [{"М": i + 1, ent: r["name"]} for i, (_, r) in enumerate(members.iterrows())]
                                df_heat = pd.DataFrame(tdata)
                                current_row = write_df(ws_stage, df_heat, current_row)

                    current_row += 2  # отступ между группами

            auto_width(ws_stage)

    # ===== Лист: Итоговые результаты =====
    if str(tourn["status"]) == "finished":
        ws_overall = wb.create_sheet("Итоги")
        overall = compute_overall_standings(tournament_id)
        if not overall.empty:
            entity = "Команда" if is_team_export else "Пилот"
            ov_data = []
            for _, row in overall.iterrows():
                ov_data.append({
                    "Место": int(row["place"]),
                    entity: row["name"],
                    "Этап выбывания": row["stage"],
                    "Результат": row["detail"],
                })
            df_ov = pd.DataFrame(ov_data)
            ws_overall.cell(row=1, column=1, value="ИТОГОВЫЕ РЕЗУЛЬТАТЫ ТУРНИРА").font = Font(bold=True, size=13)
            write_df(ws_overall, df_ov, start_row=3, medal_col=True)
        auto_width(ws_overall)

    # Записываем в буфер
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def format_time(seconds: Optional[float]) -> str:
    """Форматирует секунды в мм:сс.мс"""
    if seconds is None or (isinstance(seconds, float) and math.isnan(seconds)):
        return "—"
    m = int(seconds) // 60
    s = seconds - m * 60
    return f"{m}:{s:06.3f}"


def parse_time(time_str: str) -> Optional[float]:
    """Парсит время из строки (поддержка форматов: 90.5, 1:30.5)"""
    if not time_str or time_str.strip() == "":
        return None
    time_str = time_str.strip().replace(",", ".")
    if ":" in time_str:
        parts = time_str.split(":")
        try:
            return float(parts[0]) * 60 + float(parts[1])
        except ValueError:
            return None
    try:
        return float(time_str)
    except ValueError:
        return None


# ============================================================
# ПРИЛОЖЕНИЕ
# ============================================================

st.set_page_config(
    page_title="Гонки дронов (БВС) — Турнир",
    page_icon="🏁",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# ЗАЩИТА ПАРОЛЕМ
# ============================================================
APP_PASSWORD = st.secrets.get("APP_PASSWORD", os.environ.get("APP_PASSWORD", ""))


def check_password():
    if not APP_PASSWORD:
        st.error("⚠️ Пароль не настроен! Добавьте APP_PASSWORD в Streamlit Secrets.")
        st.stop()
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
    if st.session_state["authenticated"]:
        return True
    st.markdown("## 🔐 Вход в систему")
    st.markdown("Для доступа к системе управления турнирами введите пароль.")
    password = st.text_input("Пароль", type="password", key="password_input")
    if st.button("Войти", type="primary"):
        if password == APP_PASSWORD:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("❌ Неверный пароль")
    return False


if not check_password():
    st.stop()

init_db()
st.markdown(BASE_CSS, unsafe_allow_html=True)

# --- Сайдбар ---
with st.sidebar:
    lang_options = ["RU", "EN"]
    lang_idx = lang_options.index(st.session_state.get("lang", "RU")) if st.session_state.get("lang", "RU") in lang_options else 0
    lang = st.selectbox("🌐 " + I18N["RU"]["language"], lang_options, index=lang_idx, key="lang")
    st.divider()
    if st.button("🚪 Выйти", use_container_width=True):
        st.session_state["authenticated"] = False
        st.rerun()
    st.divider()

    st.header("🏁 " + T("tournament"))
    if st.session_state.get("tournament_just_created"):
        st.session_state["tournament_just_created"] = False
        if "tournament_selectbox" in st.session_state:
            del st.session_state["tournament_selectbox"]
    tdf = qdf("SELECT * FROM tournaments ORDER BY id DESC")
    t_map = {f'{r["name"]}': int(r["id"]) for _, r in tdf.iterrows()} if not tdf.empty else {}
    id_to_name = {v: k for k, v in t_map.items()}
    options = [T("create_new")] + list(t_map.keys())

    # Инициализация: после создания — tournament_select_init; при первом заходе — selected_tournament
    # Используем index только при инициализации, иначе виджет хранит выбор в key
    if "tournament_select_init" in st.session_state:
        if "tournament_selectbox" in st.session_state:
            del st.session_state["tournament_selectbox"]
        init_id = int(st.session_state.pop("tournament_select_init"))
        init_name = id_to_name.get(init_id) if init_id in id_to_name else None
        default_idx = options.index(init_name) if init_name and init_name in options else 0
        sel = st.selectbox(T("select_tournament"), options, index=default_idx, key="tournament_selectbox")
    elif "selected_tournament" in st.session_state and "tournament_selectbox" not in st.session_state:
        saved_id = st.session_state["selected_tournament"]
        if saved_id in id_to_name and id_to_name[saved_id] in options:
            default_idx = options.index(id_to_name[saved_id])
            sel = st.selectbox(T("select_tournament"), options, index=default_idx, key="tournament_selectbox")
        else:
            sel = st.selectbox(T("select_tournament"), options, key="tournament_selectbox")
    else:
        sel = st.selectbox(T("select_tournament"), options, key="tournament_selectbox")

    DISCIPLINES = {
        "drone_individual": T("drone_individual"),
        "sim_individual": T("sim_individual"),
        "sim_team": T("sim_team"),
    }

    if sel == T("create_new"):
        if "selected_tournament" in st.session_state:
            del st.session_state["selected_tournament"]
        st.subheader(T("create_new_header"))
        name = st.text_input(T("tournament_name"), value=f"Турнир {datetime.now().strftime('%d.%m.%Y')}")
        disc_key = st.selectbox(T("discipline"), list(DISCIPLINES.keys()),
                                format_func=lambda k: DISCIPLINES[k])

        # Условные дефолты в зависимости от дисциплины
        if disc_key in ("sim_individual", "sim_team"):
            default_time = 120.0
            default_laps = 3
        else:
            default_time = 90.0
            default_laps = 3

        time_limit = st.number_input(T("time_limit"), value=default_time, min_value=10.0, step=5.0)
        total_laps = st.number_input(T("total_laps"), value=default_laps, min_value=1, step=1)

        qual_attempts_val = 1
        if disc_key == "drone_individual":
            qual_attempts_val = st.number_input(
                "Попыток в квалификации (засчитывается лучший результат)",
                min_value=1, max_value=5, value=2, step=1,
                help="От 1 до 5. В квалификации участник делает указанное число вылетов, засчитывается лучший.")

        # Для симулятора — автоматически sum_all
        scoring_mode_val = "sum_all" if disc_key in ("sim_individual", "sim_team") else "none"
        if disc_key in ("sim_individual", "sim_team"):
            st.caption("📊 Подсчёт: сумма очков за 6 вылетов (2 трассы × 3 попытки). "
                       "Очки: 1м=4, 2м=3, 3м=2, 4м=1, DNF=0. Макс 24 очка.")
        if disc_key == "sim_team":
            st.caption("👥 Командный зачёт: 2 пилота в команде. Время за попытку суммируется.")

        if st.button(T("create_tournament"), type="primary"):
            if not name.strip():
                st.error("Введите название турнира!")
                st.stop()
            exec_sql("""INSERT INTO tournaments(name, discipline, time_limit_seconds, total_laps, scoring_mode,
                        qual_attempts, status, created_at)
                        VALUES(?,?,?,?,?,?,?,?)""",
                     (name.strip(), disc_key, time_limit, int(total_laps), scoring_mode_val,
                      int(qual_attempts_val), "setup",
                      datetime.now().isoformat(timespec="seconds")))
            new_id = int(qdf("SELECT id FROM tournaments ORDER BY id DESC LIMIT 1").iloc[0]["id"])
            st.session_state["selected_tournament"] = new_id
            st.session_state["tournament_select_init"] = new_id
            st.session_state["tournament_just_created"] = True
            st.rerun()
        tournament_id = None
    else:
        tournament_id = t_map[sel]
        st.session_state["selected_tournament"] = tournament_id  # сохраняем для rerun после действий

        # Редактирование названия турнира
        rename_key = "rename_tournament_mode"
        if st.session_state.get(rename_key, False):
            new_t_name = st.text_input("Новое название", value=sel, key="rename_input")
            rc1, rc2 = st.columns(2)
            with rc1:
                if st.button("✅ Сохранить", use_container_width=True, key="rename_save"):
                    if new_t_name.strip():
                        exec_sql("UPDATE tournaments SET name=? WHERE id=?", (new_t_name.strip(), tournament_id))
                        st.session_state[rename_key] = False
                        st.session_state["selected_tournament"] = tournament_id
                        st.rerun()
                    else:
                        st.error("Название не может быть пустым!")
            with rc2:
                if st.button("❌ Отмена", use_container_width=True, key="rename_cancel"):
                    st.session_state[rename_key] = False
                    st.rerun()
        else:
            if st.button("✏️ Переименовать", use_container_width=True, key="rename_btn"):
                st.session_state[rename_key] = True
                st.rerun()

    # --- Управление базой данных ---
    st.divider()
    with st.expander("🗄️ Управление БД", expanded=False):
        # Экспорт БД
        st.markdown("**📤 Экспорт базы данных**")
        st.caption("Скачайте полную копию БД со всеми турнирами")
        if os.path.exists(DB_PATH):
            with open(DB_PATH, "rb") as f:
                db_bytes = f.read()
            st.download_button("📥 Скачать БД (.db)", data=db_bytes,
                               file_name="tournament_backup.db",
                               mime="application/octet-stream",
                               use_container_width=True)

        st.divider()

        # Импорт БД
        st.markdown("**📥 Импорт базы данных**")
        st.caption("⚠️ Загрузка заменит ВСЮ текущую базу данных!")
        uploaded = st.file_uploader("Выберите .db файл", type=["db"], key="db_upload")
        if uploaded is not None:
            if not st.session_state.get("confirm_db_import", False):
                if st.button("⚠️ Заменить текущую БД", type="primary", use_container_width=True):
                    st.session_state["confirm_db_import"] = True
                    st.rerun()
            else:
                st.warning("**Вы уверены?** Все текущие данные будут заменены!")
                ic1, ic2 = st.columns(2)
                with ic1:
                    if st.button("✅ Да, заменить", type="primary", use_container_width=True):
                        # Закрываем все соединения перед перезаписью
                        try:
                            conn = sqlite3.connect(DB_PATH)
                            conn.close()
                        except Exception:
                            pass
                        with open(DB_PATH, "wb") as f:
                            f.write(uploaded.getvalue())
                        # Проверяем целостность новой БД
                        try:
                            test_conn = sqlite3.connect(DB_PATH)
                            test_conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
                            test_conn.close()
                        except Exception:
                            st.error("⚠️ Загруженный файл не является корректной базой данных!")
                            st.stop()
                        st.session_state["confirm_db_import"] = False
                        st.success("✅ БД успешно импортирована!")
                        st.rerun()
                with ic2:
                    if st.button("❌ Отмена", use_container_width=True):
                        st.session_state["confirm_db_import"] = False
                        st.rerun()

    # --- Удаление турнира ---
    if tournament_id is not None:
        st.divider()
        with st.expander("🗑️ Удалить турнир", expanded=False):
            st.warning(f"Удаление турнира **{sel}** безвозвратно!")
            del_key = "confirm_delete_tournament"
            if not st.session_state.get(del_key, False):
                if st.button("🗑️ Удалить этот турнир", type="primary", use_container_width=True):
                    st.session_state[del_key] = True
                    st.rerun()
            else:
                st.error("⚠️ Вы уверены? Это действие нельзя отменить!")
                dc1, dc2 = st.columns(2)
                with dc1:
                    if st.button("✅ Да, удалить", type="primary", use_container_width=True):
                        # Удаляем вручную для надёжности (на случай если FK не работает)
                        stage_ids = qdf("SELECT id FROM stages WHERE tournament_id=?", (tournament_id,))
                        for _, sr in stage_ids.iterrows():
                            sid = int(sr["id"])
                            group_ids = qdf("SELECT id FROM groups WHERE stage_id=?", (sid,))
                            for _, gr in group_ids.iterrows():
                                gid = int(gr["id"])
                                heat_ids = qdf("SELECT id FROM heats WHERE group_id=?", (gid,))
                                for _, hr in heat_ids.iterrows():
                                    exec_sql("DELETE FROM heat_results WHERE heat_id=?", (int(hr["id"]),))
                                exec_sql("DELETE FROM heats WHERE group_id=?", (gid,))
                                exec_sql("DELETE FROM group_members WHERE group_id=?", (gid,))
                            exec_sql("DELETE FROM groups WHERE stage_id=?", (sid,))
                        exec_sql("DELETE FROM stages WHERE tournament_id=?", (tournament_id,))
                        p_ids = qdf("SELECT id FROM participants WHERE tournament_id=?", (tournament_id,))
                        for _, pr in p_ids.iterrows():
                            exec_sql("DELETE FROM team_pilots WHERE participant_id=?", (int(pr["id"]),))
                        exec_sql("DELETE FROM qualification_results WHERE tournament_id=?", (tournament_id,))
                        exec_sql("DELETE FROM participants WHERE tournament_id=?", (tournament_id,))
                        exec_sql("DELETE FROM tournaments WHERE id=?", (tournament_id,))
                        st.session_state[del_key] = False
                        if "selected_tournament" in st.session_state:
                            del st.session_state["selected_tournament"]
                        if "tournament_selectbox" in st.session_state:
                            del st.session_state["tournament_selectbox"]
                        st.success("✅ Турнир удалён!")
                        st.rerun()
                with dc2:
                    if st.button("❌ Отмена", use_container_width=True):
                        st.session_state[del_key] = False
                        st.rerun()

if tournament_id is None:
    st.header(T("app_title"))
    st.info("Выберите турнир или создайте новый.")
    st.stop()

# --- Данные турнира ---
tourn = get_tournament(tournament_id)
if tourn is None:
    st.error("Турнир не найден в базе данных.")
    st.stop()
discipline = str(tourn["discipline"])
t_status = str(tourn["status"])
time_limit = float(tourn["time_limit_seconds"])
total_laps = int(tourn["total_laps"])
qual_attempts = int(tourn.get("qual_attempts", 1))
scoring_mode = str(tourn.get("scoring_mode", "none"))
p_count = participant_count(tournament_id)
is_sim = discipline in ("sim_individual", "sim_team")
is_team = discipline == "sim_team"

with st.sidebar:
    st.caption(f"📋 {DISCIPLINES.get(discipline, discipline)}")
    st.caption(f"⏱️ {time_limit}с / {total_laps} кр.")
    if is_sim:
        st.caption("📊 Сумма очков за 6 вылетов")
    if is_team:
        st.caption(f"👥 {p_count} команд")
    else:
        st.caption(f"👥 {p_count} участников")

# Все три дисциплины поддерживаются

st.title(str(tourn["name"]))

# --- Навигация ---
tabs = st.tabs([
    T("nav_overview"),
    T("nav_participants"),
    T("nav_qualification"),
    T("nav_bracket"),
    T("nav_playoff"),
    T("nav_final"),
    "📋 Результаты",
])

# ============================================================
# TAB 0: Обзор
# ============================================================
with tabs[0]:
    st.subheader(T("overview_title"))
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Команд" if is_team else T("total_participants"), p_count)
    with c2:
        status_labels = {"setup": T("status_setup"), "qualification": T("status_qualification"),
                         "bracket": T("status_bracket"), "finished": T("status_finished")}
        st.metric(T("tournament_status"), status_labels.get(t_status, t_status))
    with c3:
        bracket = get_bracket_for_tournament(tournament_id)
        all_stages = get_all_stages(tournament_id)
        done_count = len(all_stages[all_stages["status"] == "done"]) if not all_stages.empty else 0
        total_stages = len(bracket) if bracket else 0
        st.metric("Этапов", f"{done_count} / {total_stages}")

    # Прогресс
    if bracket:
        st.markdown("**Прогресс турнира:**")
        qual_css = "completed" if t_status not in ("setup",) else ("active" if t_status == "qualification" else "pending")
        if t_status == "setup":
            qual_css = "pending"
        elif t_status == "qualification":
            qual_css = "active"
        else:
            qual_css = "completed"

        progress_html = '<div class="tournament-progress">'
        qual_dot = "✓" if qual_css == "completed" else ""
        progress_html += f'<div class="progress-step {qual_css}">'
        progress_html += f'<div class="progress-dot">{qual_dot}</div>'
        progress_html += '<div class="progress-label">Квалификация</div></div>'

        for idx, sd in enumerate(bracket):
            stage_row = all_stages[all_stages["stage_idx"] == idx] if not all_stages.empty else pd.DataFrame()
            if not stage_row.empty:
                s = stage_row.iloc[0]["status"]
                css = "completed" if s == "done" else "active"
            else:
                css = "pending"
            sname = sd.display_name.get(lang, sd.code)
            dot_icon = "✓" if css == "completed" else ""
            progress_html += f'<div class="progress-step {css}">'
            progress_html += f'<div class="progress-dot">{dot_icon}</div>'
            progress_html += f'<div class="progress-label">{sname}</div></div>'
        progress_html += '</div>'
        st.markdown(progress_html, unsafe_allow_html=True)

    # --- Экспорт турнира в Excel ---
    if t_status != "setup":
        st.divider()
        st.markdown("### 📥 Экспорт турнира")
        st.caption("Скачайте полный отчёт по турниру в формате Excel (все этапы, результаты, сводки)")
        try:
            with st.spinner("Генерация отчёта..."):
                excel_data = export_tournament_excel(tournament_id)
            safe_name = str(tourn["name"]).replace(" ", "_").replace("/", "-")[:30]
            st.download_button(
                label="📥 Скачать полный отчёт (Excel)",
                data=excel_data,
                file_name=f"Турнир_{safe_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Ошибка при генерации отчёта: {e}")

# ============================================================
# TAB 1: Участники
# ============================================================
with tabs[1]:
    st.subheader(T("participants_title"))

    col1, col2 = st.columns([1, 2])

    with col1:
        st.markdown(f"### {'Добавить команду' if is_team else T('add_participant')}")
        if t_status in ("bracket", "finished"):
            st.info("🔒 Добавление участников недоступно — турнир уже продвинулся дальше квалификации.")
        else:
            if is_team:
                with st.form("add_team", clear_on_submit=True):
                    team_name = st.text_input("Название команды")
                    pilot1 = st.text_input("Пилот 1")
                    pilot2 = st.text_input("Пилот 2")
                    if st.form_submit_button(T("add"), type="primary"):
                        if team_name.strip() and pilot1.strip() and pilot2.strip():
                            exec_sql("INSERT INTO participants(tournament_id, name) VALUES(?,?)",
                                     (tournament_id, team_name.strip()))
                            new_pid = int(qdf("SELECT id FROM participants WHERE tournament_id=? ORDER BY id DESC LIMIT 1",
                                              (tournament_id,)).iloc[0]["id"])
                            exec_sql("INSERT INTO team_pilots(participant_id, pilot1_name, pilot2_name) VALUES(?,?,?)",
                                     (new_pid, pilot1.strip(), pilot2.strip()))
                            st.success(T("saved"))
                            st.rerun()
                        else:
                            st.warning("Заполните все три поля: название команды и имена обоих пилотов.")
            else:
                with st.form("add_pilot", clear_on_submit=True):
                    pname = st.text_input(T("pilot_name"))
                    if st.form_submit_button(T("add"), type="primary"):
                        if pname.strip():
                            exec_sql("INSERT INTO participants(tournament_id, name) VALUES(?,?)",
                                     (tournament_id, pname.strip()))
                            st.success(T("saved"))
                            st.rerun()

        st.divider()
        st.markdown(f"### {T('random_draw')}")

        # Проверяем, была ли уже проведена жеребьёвка
        has_numbers = int(qdf("SELECT COUNT(*) as c FROM participants WHERE tournament_id=? AND start_number IS NOT NULL",
                              (tournament_id,)).iloc[0]["c"])
        draw_done = has_numbers > 0

        if not draw_done:
            # Первая жеребьёвка — простая кнопка
            if st.button(T("random_draw"), type="primary"):
                pdf = qdf("SELECT id FROM participants WHERE tournament_id=?", (tournament_id,))
                if pdf.empty:
                    st.warning("Нет команд" if is_team else "Нет участников")
                else:
                    ids = pdf["id"].tolist()
                    random.shuffle(ids)
                    for idx, pid in enumerate(ids):
                        exec_sql("UPDATE participants SET start_number=? WHERE id=?", (idx + 1, pid))
                    exec_sql("UPDATE tournaments SET status='qualification' WHERE id=?", (tournament_id,))
                    st.success(T("draw_done"))
                    st.balloons()
                    st.rerun()
        else:
            st.success("✅ Жеребьёвка проведена")
            if t_status in ("bracket", "finished"):
                # Квалификация уже завершена — пережеребьёвка запрещена
                st.info("🔒 Пережеребьёвка недоступна — квалификация уже завершена и турнир продвинулся дальше.")
            else:
                # Повторная жеребьёвка — с двойным подтверждением
                redraw_key = "confirm_redraw"
                if not st.session_state.get(redraw_key, False):
                    if st.button("🔄 Провести жеребьёвку ещё раз"):
                        st.session_state[redraw_key] = True
                        st.rerun()
                else:
                    st.warning("⚠️ Вы уверены? Все текущие стартовые номера будут перемешаны заново.")
                    cc1, cc2 = st.columns(2)
                    with cc1:
                        if st.button("✅ Да, перемешать", type="primary", use_container_width=True):
                            pdf = qdf("SELECT id FROM participants WHERE tournament_id=?", (tournament_id,))
                            ids = pdf["id"].tolist()
                            random.shuffle(ids)
                            for idx, pid in enumerate(ids):
                                exec_sql("UPDATE participants SET start_number=? WHERE id=?", (idx + 1, pid))
                            st.session_state[redraw_key] = False
                            st.success(T("draw_done"))
                            st.balloons()
                            st.rerun()
                    with cc2:
                        if st.button("❌ Отмена", use_container_width=True):
                            st.session_state[redraw_key] = False
                            st.rerun()

        # --- Инструменты разработчика (скрыты) ---
        st.divider()
        with st.expander("🛠️ Инструменты разработчика", expanded=False):
            st.caption(T("demo_hint"))
            n_demo = st.number_input(T("demo_count"), min_value=4, max_value=64, value=16, step=1)
            prefix = st.text_input(T("demo_prefix"), value="Команда" if is_team else "Пилот")
            if st.button(T("demo_add")):
                if participant_count(tournament_id) > 0:
                    st.warning(T("demo_already"))
                else:
                    if is_team:
                        for i in range(1, int(n_demo) + 1):
                            exec_sql("INSERT INTO participants(tournament_id, name) VALUES(?,?)",
                                     (tournament_id, f"{prefix} {i}"))
                            new_pid = int(qdf("SELECT id FROM participants WHERE tournament_id=? ORDER BY id DESC LIMIT 1",
                                              (tournament_id,)).iloc[0]["id"])
                            exec_sql("INSERT INTO team_pilots(participant_id, pilot1_name, pilot2_name) VALUES(?,?,?)",
                                     (new_pid, f"Пилот {i}A", f"Пилот {i}B"))
                    else:
                        rows = [(tournament_id, f"{prefix} {i}") for i in range(1, int(n_demo) + 1)]
                        exec_many("INSERT INTO participants(tournament_id, name) VALUES(?,?)", rows)
                    st.success(f'{T("demo_added")}: {n_demo}')
                    st.rerun()

            st.divider()
            st.markdown("**📥 Импорт участников из Excel**")
            st.caption("⚠️ Тестовая функция")
            if is_team:
                st.caption("Импорт для командного зачёта пока не поддерживается.")
            else:
                st.caption("Формат: колонка ФИО, Категория, колонки «75 ЛЗ» (дроны) и «ТС ЛЗ» (симулятор) с «+». "
                          "Если файл в формате .xls — сохраните в Excel как .xlsx.")
                excel_upload = st.file_uploader("Файл .xlsx", type=["xlsx"], key="excel_import")
                discipline_filter = st.selectbox(
                    "Дисциплина в файле",
                    ["75 ЛЗ (дроны)", "ТС ЛЗ (симулятор)"],
                    index=0 if discipline == "drone_individual" else 1,
                    key="excel_discipline"
                )
                category_filter = st.selectbox(
                    "Категория",
                    ["Все категории", "Мальчики", "Юниорки", "Юниоры", "Девочки"],
                    key="excel_category"
                )
                if excel_upload is not None:
                    if st.button("📥 Импортировать из Excel", key="excel_import_btn"):
                        try:
                            engine = "openpyxl"
                            df = pd.read_excel(excel_upload, engine=engine, header=None)
                            if df.empty:
                                st.warning("Файл пустой.")
                            else:
                                cat_val = None if category_filter == "Все категории" else category_filter
                                disc_for_import = "drone_individual" if "75" in discipline_filter else "sim_individual"
                                names, detected = _parse_excel_discipline_list(df, disc_for_import, category_filter=cat_val)
                                added = 0
                                if not names and detected.get("75лз") is None and detected.get("тслз") is None:
                                    st.warning("Не найдены колонки «75 ЛЗ» или «ТС ЛЗ». Проверьте структуру файла.")
                                elif not names:
                                    disc_label = "75 ЛЗ" if "75" in discipline_filter else "ТС ЛЗ"
                                    st.warning(f"Нет участников с «+» в колонке {disc_label}.")
                                else:
                                    for name in names:
                                        exec_sql("INSERT INTO participants(tournament_id, name) VALUES(?,?)",
                                                 (tournament_id, name))
                                        added += 1
                                if added > 0:
                                    st.session_state["selected_tournament"] = tournament_id
                                    st.session_state["tournament_select_init"] = tournament_id
                                    st.session_state["tournament_just_created"] = True
                                    st.success(f"Импортировано: {added}")
                                    st.rerun()
                        except Exception as e:
                            st.error(f"Ошибка: {e}")

    with col2:
        participants_raw = qdf("""SELECT id, start_number, name, COALESCE(disqualified,0) as disqualified
                                  FROM participants WHERE tournament_id=?
                                  ORDER BY COALESCE(start_number, 9999), name""", (tournament_id,))

        # Для командного зачёта загрузим данные пилотов
        team_pilots_map = {}
        if is_team and not participants_raw.empty:
            tp_df = qdf("SELECT participant_id, pilot1_name, pilot2_name FROM team_pilots WHERE participant_id IN ({})".format(
                ",".join(str(int(x)) for x in participants_raw["id"].tolist())))
            for _, tpr in tp_df.iterrows():
                team_pilots_map[int(tpr["participant_id"])] = (tpr["pilot1_name"], tpr["pilot2_name"])

        if participants_raw.empty:
            st.info("Пока нет участников. Добавьте слева." if not is_team else "Пока нет команд. Добавьте слева.")
        else:
            count_label = f"**Всего: {len(participants_raw)} {'команд' if is_team else 'участников'}**"
            st.markdown(count_label)

            for _, row in participants_raw.iterrows():
                pid = int(row["id"])
                pname = row["name"]
                sn = f"#{int(row['start_number'])}" if pd.notna(row["start_number"]) else ""
                pilots = team_pilots_map.get(pid, None)
                is_dsq = bool(int(row.get("disqualified", 0)))
                dsq_suffix = f" — **{T('disqualified_full')}**" if is_dsq else ""

                locked = t_status in ("bracket", "finished")
                with st.container(border=True):
                    if locked:
                        c1, c2, c3 = st.columns([1, 4, 1])
                        with c1:
                            st.markdown(f"**{sn}**" if sn else "—")
                        with c2:
                            if is_team and pilots:
                                st.markdown(f"{pname} ({pilots[0]}, {pilots[1]}){dsq_suffix}")
                            else:
                                st.markdown(f"{pname}{dsq_suffix}")
                        with c3:
                            if st.button("🚫" if is_dsq else "⚠️", key=f"dsq_{pid}", use_container_width=True,
                                         help=T("disqualify_undo") if is_dsq else T("disqualify")):
                                set_participant_disqualified(pid, not is_dsq)
                                st.rerun()
                    else:
                        c1, c2, c3 = st.columns([1, 5, 2])
                        with c1:
                            st.markdown(f"**{sn}**" if sn else "—")
                        with c2:
                            # Inline edit
                            edit_key = f"edit_mode_{pid}"
                            if st.session_state.get(edit_key, False):
                                if is_team:
                                    new_name = st.text_input("Команда", value=pname, key=f"edit_name_{pid}")
                                    p1_val = pilots[0] if pilots else ""
                                    p2_val = pilots[1] if pilots else ""
                                    new_p1 = st.text_input("Пилот 1", value=p1_val, key=f"edit_p1_{pid}")
                                    new_p2 = st.text_input("Пилот 2", value=p2_val, key=f"edit_p2_{pid}")
                                else:
                                    new_name = st.text_input("Имя", value=pname, key=f"edit_name_{pid}", label_visibility="collapsed")
                                ec1, ec2 = st.columns(2)
                                with ec1:
                                    if st.button("✅", key=f"save_edit_{pid}", use_container_width=True):
                                        if new_name.strip():
                                            exec_sql("UPDATE participants SET name=? WHERE id=?", (new_name.strip(), pid))
                                        if is_team:
                                            p1_save = new_p1.strip() if new_p1.strip() else p1_val
                                            p2_save = new_p2.strip() if new_p2.strip() else p2_val
                                            if pilots:
                                                exec_sql("UPDATE team_pilots SET pilot1_name=?, pilot2_name=? WHERE participant_id=?",
                                                         (p1_save, p2_save, pid))
                                            else:
                                                exec_sql("INSERT INTO team_pilots(participant_id, pilot1_name, pilot2_name) VALUES(?,?,?)",
                                                         (pid, p1_save, p2_save))
                                        st.session_state[edit_key] = False
                                        st.rerun()
                                with ec2:
                                    if st.button("❌", key=f"cancel_edit_{pid}", use_container_width=True):
                                        st.session_state[edit_key] = False
                                        st.rerun()
                            else:
                                if is_team and pilots:
                                    st.markdown(f"{pname} ({pilots[0]}, {pilots[1]}){dsq_suffix}")
                                else:
                                    st.markdown(f"{pname}{dsq_suffix}")
                        with c3:
                            bc1, bc2, bc3 = st.columns(3)
                            with bc1:
                                if st.button("✏️", key=f"btn_edit_{pid}", use_container_width=True):
                                    st.session_state[f"edit_mode_{pid}"] = True
                                    st.rerun()
                            with bc2:
                                if st.button("🗑️", key=f"btn_del_{pid}", use_container_width=True):
                                    exec_sql("DELETE FROM team_pilots WHERE participant_id=?", (pid,))
                                    exec_sql("DELETE FROM participants WHERE id=?", (pid,))
                                    exec_sql("DELETE FROM qualification_results WHERE participant_id=?", (pid,))
                                    st.rerun()
                            with bc3:
                                if st.button("🚫" if is_dsq else "⚠️", key=f"dsq_{pid}", use_container_width=True,
                                             help=T("disqualify_undo") if is_dsq else T("disqualify")):
                                    set_participant_disqualified(pid, not is_dsq)
                                    st.rerun()

# ============================================================
# TAB 2: Квалификация
# ============================================================
with tabs[2]:
    st.subheader(T("qual_title"))

    if t_status == "setup":
        st.info("Сначала добавьте участников и проведите жеребьёвку на вкладке 'Участники'")
    elif t_status in ("bracket", "finished"):
        st.success("✅ Квалификация завершена!")
        ranking = get_qual_ranking(tournament_id)
        if not ranking.empty:
            advancing = compute_bracket_size(len(ranking))
            st.info(T("qual_cutoff").format(advancing, len(ranking)))
            pilot_label = "Команда" if is_team else T("pilot")
            if is_sim:
                display = ranking[["place", "name", "start_number", "time_seconds",
                                   "laps_completed"]].copy()
                display.columns = [T("place"), pilot_label, "№", T("time_seconds"), T("laps_completed")]
            else:
                display = ranking[["place", "name", "start_number", "time_seconds",
                                   "laps_completed", "completed_all_laps", "projected_time"]].copy()
                display.columns = [T("place"), pilot_label, "№", T("time_seconds"),
                                   T("laps_completed"), T("completed_all"), T("projected_time")]
            styled = style_qual_table(display, advancing)
            st.dataframe(styled, use_container_width=True, hide_index=True)
            st.caption("🟢 Зелёный = проходит | 🔴 Красный = не проходит")
    else:
        if is_team:
            st.info("Введите результаты каждой команды (время каждого пилота суммируется). Лимит 2 минуты, 3 круга.")
        elif is_sim:
            st.info(T("sim_qual_info"))
        else:
            st.info(T("qual_info"))
        if qual_attempts > 1:
            st.caption(f"⏱️ Лимит: {time_limit} сек | 🔄 Кругов: {total_laps} | 📋 {qual_attempts} попыток (засчитывается лучший результат)")
        else:
            st.caption(f"⏱️ Лимит: {time_limit} сек | 🔄 Кругов: {total_laps}")

        all_participants = qdf("""
            SELECT p.id as pid, p.name, p.start_number, COALESCE(p.disqualified,0) as disqualified,
                   (SELECT COUNT(*) FROM qualification_results qr
                    WHERE qr.participant_id=p.id AND qr.tournament_id=p.tournament_id
                    AND qr.time_seconds IS NOT NULL) as attempts_filled
            FROM participants p
            WHERE p.tournament_id=? AND p.start_number IS NOT NULL
            ORDER BY p.start_number
        """, (tournament_id,))

        if all_participants.empty:
            st.warning("Проведите жеребьёвку на вкладке 'Участники'")
        else:
            st.markdown("### Ввод результатов")

            # Для командного зачёта загружаем пилотов
            qual_team_map = {}
            if is_team:
                all_pids = [int(r["pid"]) for _, r in all_participants.iterrows()]
                if all_pids:
                    tp_q = qdf("SELECT participant_id, pilot1_name, pilot2_name FROM team_pilots WHERE participant_id IN ({})".format(
                        ",".join(str(x) for x in all_pids)))
                    for _, tpr in tp_q.iterrows():
                        qual_team_map[int(tpr["participant_id"])] = (tpr["pilot1_name"], tpr["pilot2_name"])

            qual_dsq_pids = get_disqualified_pids(tournament_id)
            for _, row in all_participants.iterrows():
                pid = int(row["pid"])
                sn = int(row["start_number"])
                name = row["name"]
                is_dsq_qual = pid in qual_dsq_pids
                q_pilots = qual_team_map.get(pid, None) if is_team else None

                attempts_filled = int(row.get("attempts_filled", 0))
                expander_label = f"**#{sn} {name}**"
                if is_team and q_pilots:
                    expander_label += f" ({q_pilots[0]}, {q_pilots[1]})"
                if is_dsq_qual:
                    expander_label += f" — **{T('disqualified')}**"
                else:
                    if qual_attempts > 1:
                        expander_label += f" {attempts_filled}/{qual_attempts} ✅" if attempts_filled >= qual_attempts else f" {attempts_filled}/{qual_attempts} ⏳"
                    else:
                        expander_label += " ✅" if attempts_filled >= 1 else " ⏳"

                with st.expander(expander_label, expanded=(attempts_filled < qual_attempts) and not is_dsq_qual):
                    if is_dsq_qual:
                        st.caption(f"🚫 {T('disqualified_full')} — результат проставлен автоматически")
                        continue
                    if is_team:
                        # Командный зачёт: два времени пилотов, автосумма
                        atts_t = get_participant_qual_attempts(tournament_id, pid)
                        att_row_t = atts_t.iloc[0] if not atts_t.empty else {}
                        existing_time = float(att_row_t.get("time_seconds", 0)) if att_row_t.get("time_seconds") else 0.0
                        existing_laps_t = float(att_row_t.get("laps_completed", 0)) if att_row_t.get("laps_completed") else 0.0
                        p1_label = q_pilots[0] if q_pilots else "Пилот 1"
                        p2_label = q_pilots[1] if q_pilots else "Пилот 2"
                        c1, c2, c3 = st.columns([2, 2, 2])
                        with c1:
                            t1_val = st.number_input(
                                f"⏱️ {p1_label} (сек)", min_value=0.0, max_value=999.0,
                                value=_safe_time_for_input(existing_time) / 2,
                                step=0.001, key=f"qt1_{pid}", format="%.3f")
                        with c2:
                            t2_val = st.number_input(
                                f"⏱️ {p2_label} (сек)", min_value=0.0, max_value=999.0,
                                value=_safe_time_for_input(existing_time) / 2,
                                step=0.001, key=f"qt2_{pid}", format="%.3f")
                        with c3:
                            sum_time = t1_val + t2_val
                            st.metric("Сумма (сек)", f"{sum_time:.3f}")

                        laps_col1, _ = st.columns([2, 4])
                        with laps_col1:
                            laps_val = st.number_input(
                                "Круги.Препятствия", min_value=0.0, max_value=99.0,
                                value=existing_laps_t, step=0.1, key=f"ql_{pid}", format="%.1f")
                        all_laps = laps_val >= total_laps

                        if st.button("💾 Сохранить", key=f"qs_{pid}"):
                            if sum_time > 0:
                                save_qual_result(tournament_id, pid, sum_time, laps_val, all_laps, total_laps)
                                st.success(T("saved"))
                                st.rerun()
                            else:
                                st.error("Введите время обоих пилотов!")
                    elif is_sim:
                        # Симулятор личный: только время и круги, без расчётного
                        atts_s = get_participant_qual_attempts(tournament_id, pid)
                        att_row_s = atts_s.iloc[0] if not atts_s.empty else {}
                        existing_time_s = float(att_row_s.get("time_seconds", 0)) if att_row_s.get("time_seconds") else 0.0
                        existing_laps_s = float(att_row_s.get("laps_completed", 0)) if att_row_s.get("laps_completed") else 0.0
                        c1, c2 = st.columns([2, 2])
                        with c1:
                            time_val = st.number_input(
                                f"Время (сек)", min_value=0.0, max_value=999.0,
                                value=_safe_time_for_input(existing_time_s), step=0.001, key=f"qt_{pid}", format="%.3f")
                        with c2:
                            laps_val = st.number_input(
                                "Круги.Препятствия", min_value=0.0, max_value=99.0,
                                value=existing_laps_s, step=0.1, key=f"ql_{pid}", format="%.1f")
                        all_laps = laps_val >= total_laps  # Авто-определяем

                        if st.button("💾 Сохранить", key=f"qs_{pid}"):
                            if time_val > 0:
                                save_qual_result(tournament_id, pid, time_val, laps_val, all_laps, total_laps)
                                st.success(T("saved"))
                                st.rerun()
                            else:
                                st.error("Введите время!")
                    else:
                        # Дроны: полный набор полей (может быть несколько попыток)
                        attempts_data = get_participant_qual_attempts(tournament_id, pid)
                        attempts_map = {int(r["attempt_no"]): r for _, r in attempts_data.iterrows()} if not attempts_data.empty else {}

                        for att_no in range(1, qual_attempts + 1):
                            att_row = attempts_map.get(att_no, {})
                            ex_time = float(att_row["time_seconds"]) if att_row.get("time_seconds") else 0.0
                            ex_laps = float(att_row["laps_completed"]) if att_row.get("laps_completed") else 0.0
                            ex_time = _safe_time_for_input(ex_time)
                            ex_laps = 0.0 if (isinstance(ex_laps, float) and math.isnan(ex_laps)) else max(0.0, min(99.0, ex_laps))
                            ex_all = bool(int(att_row.get("completed_all_laps", 0))) if att_no in attempts_map else False

                            st.markdown(f"**Вылет {att_no}**" + (" ✅" if att_no in attempts_map else ""))
                            c1, c2, c3, c4 = st.columns([2, 2, 1, 2])
                            with c1:
                                time_val = st.number_input(
                                    f"Время (сек)", min_value=0.0, max_value=999.0,
                                    value=ex_time, step=0.001,
                                    key=f"qt_{pid}_{att_no}", format="%.3f")
                            with c2:
                                laps_val = st.number_input(
                                    "Круги.Препятствия", min_value=0.0, max_value=99.0,
                                    value=ex_laps, step=0.1, key=f"ql_{pid}_{att_no}", format="%.1f")
                            with c3:
                                all_laps = st.checkbox("Все круги", value=ex_all, key=f"qa_{pid}_{att_no}",
                                                       help="Отметьте, если пилот прошёл все круги за отведённое время")
                            with c4:
                                if time_val > 0 and laps_val > 0:
                                    proj = time_val if all_laps else calc_projected_time(time_val, laps_val, total_laps)
                                    st.metric("Расчётное", format_time(proj))

                            if st.button(f"💾 Сохранить вылет {att_no}", key=f"qs_{pid}_{att_no}"):
                                if time_val > 0:
                                    save_qual_result(tournament_id, pid, time_val, laps_val, all_laps, total_laps, attempt_no=att_no)
                                    st.success(T("saved"))
                                    st.rerun()
                                else:
                                    st.error("Введите время!")
                            if att_no < qual_attempts:
                                st.divider()

            # --- Таблица результатов ---
            st.divider()
            st.markdown("### 📊 Текущая таблица квалификации")

            ranking = get_qual_ranking(tournament_id)
            if not ranking.empty:
                advancing = compute_bracket_size(len(ranking))
                total_p = participant_count(tournament_id)
                st.info(T("qual_cutoff").format(advancing, total_p))

                pilot_col_name = "Команда" if is_team else "Пилот"
                if is_sim:
                    display = ranking[["place", "name", "start_number", "time_seconds",
                                       "laps_completed"]].copy()
                    display.columns = ["Место", pilot_col_name, "№", "Время (сек)", "Круги"]
                else:
                    display = ranking[["place", "name", "start_number", "time_seconds",
                                       "laps_completed", "completed_all_laps", "projected_time"]].copy()
                    display.columns = ["Место", pilot_col_name, "№", "Время (сек)", "Круги", "Все 3", "Расчётное"]
                styled = style_qual_table(display, advancing)
                st.dataframe(styled, use_container_width=True, hide_index=True)
                st.caption("🟢 Зелёный = проходит | 🔴 Красный = не проходит")

                download_csv_button(display, T("download_csv"), f"qualification_{tournament_id}.csv")

                # Кнопка завершения
                st.divider()
                if qual_attempts > 1:
                    filled = int(qdf("""
                        SELECT COUNT(*) as c FROM (
                            SELECT participant_id FROM qualification_results
                            WHERE tournament_id=? AND time_seconds IS NOT NULL
                            GROUP BY participant_id
                            HAVING COUNT(*) >= ?
                        )
                    """, (tournament_id, qual_attempts)).iloc[0]["c"])
                else:
                    filled = int(qdf("SELECT COUNT(DISTINCT participant_id) as c FROM qualification_results WHERE tournament_id=? AND time_seconds IS NOT NULL",
                                    (tournament_id,)).iloc[0]["c"])
                if filled < total_p:
                    st.warning(f"Результаты введены: {filled} из {total_p}")

                if filled < total_p:
                    # Не все результаты — нужно двойное подтверждение
                    confirm_qual_key = "confirm_qual_finish"
                    if not st.session_state.get(confirm_qual_key, False):
                        if st.button(T("qual_finish"), type="primary", use_container_width=True):
                            if filled == 0:
                                st.error(T("qual_not_all"))
                            else:
                                st.session_state[confirm_qual_key] = True
                                st.rerun()
                    else:
                        st.warning(f"⚠️ Заполнено {filled} из {total_p} результатов. Продолжить?")
                        qc1, qc2 = st.columns(2)
                        with qc1:
                            if st.button("✅ Да, завершить квалификацию", type="primary", use_container_width=True):
                                st.session_state[confirm_qual_key] = False
                                start_bracket(tournament_id)
                                st.success(T("qual_done"))
                                st.balloons()
                                st.rerun()
                        with qc2:
                            if st.button("❌ Отмена", use_container_width=True):
                                st.session_state[confirm_qual_key] = False
                                st.rerun()
                else:
                    if st.button(T("qual_finish"), type="primary", use_container_width=True):
                        start_bracket(tournament_id)
                        st.success(T("qual_done"))
                        st.balloons()
                        st.rerun()
            else:
                st.info("Введите результаты выше")

# ============================================================
# TAB 3: Сетка
# ============================================================
with tabs[3]:
    st.subheader(T("bracket_title"))

    bracket = get_bracket_for_tournament(tournament_id)
    all_stages = get_all_stages(tournament_id)

    if not bracket:
        st.info(T("waiting_for_qual"))
    else:
        # Статус
        if t_status == "qualification":
            st.info("⚡ Идёт квалификация — сетка сформируется после её завершения")
        elif t_status == "bracket":
            active = get_active_stage(tournament_id)
            if active is not None:
                active_idx = int(active["stage_idx"])
                active_sd = bracket[active_idx]
                sname = active_sd.display_name.get(lang, active_sd.code)
                if active_sd.code == "F":
                    st.success(f"🏆 Идёт: **{sname}**")
                else:
                    st.success(f"🔥 Идёт: **{sname}**")
        elif t_status == "finished":
            st.success("🏆 **ТУРНИР ЗАВЕРШЁН!**")

        st.divider()

        # Построение визуальной сетки (HTML)
        bracket_html = '<div class="bracket-container">'

        for idx, sd in enumerate(bracket):
            sname = sd.display_name.get(lang, sd.code)
            stage_row = all_stages[all_stages["stage_idx"] == idx] if not all_stages.empty else pd.DataFrame()
            status = ""
            stage_id_br = None
            if not stage_row.empty:
                stage_id_br = int(stage_row.iloc[0]["id"])
                status = stage_row.iloc[0]["status"]

            # Заголовок раунда
            title_class = "bracket-round-title"
            if sd.code == "F":
                title_class += " final-round"
            elif status == "active":
                title_class += " active-round"
            elif status == "done":
                title_class += " done-round"

            status_icon = ""
            if status == "active":
                status_icon = " ▶"
            elif status == "done":
                status_icon = " ✓"

            round_icon = "🏆 " if sd.code == "F" else ""
            has_connector = idx < len(bracket) - 1

            # Обёртка: заголовок + (группы + коннектор)
            bracket_html += '<div class="bracket-round-wrapper">'
            bracket_html += f'<div class="{title_class}">{round_icon}{sname}{status_icon}</div>'
            bracket_html += '<div class="bracket-groups-row">'
            bracket_html += '<div class="bracket-groups-col">'

            # --- Содержимое групп ---
            if sd.code == "F" and stage_id_br:
                if is_sim:
                    # Финал симулятора
                    sim_fin = compute_sim_final_standings(stage_id_br, scoring_mode)
                    if not sim_fin.empty and int(sim_fin.iloc[0].get("total_points", 0)) > 0:
                        bracket_html += '<div class="bracket-group">'
                        medals_html = {1: "🥇", 2: "🥈", 3: "🥉"}
                        medal_class = {1: "gold", 2: "silver", 3: "bronze"}
                        for _, fr in sim_fin.iterrows():
                            rank = int(fr["rank"])
                            cls = medal_class.get(rank, "")
                            medal = medals_html.get(rank, f"{rank}.")
                            bracket_html += f'<div class="bracket-player {cls}">'
                            bracket_html += f'<span>{medal} {fr["name"]}</span>'
                            bracket_html += f'<span><b>{int(fr["total_points"])} оч.</b></span>'
                            bracket_html += '</div>'
                        bracket_html += '</div>'
                    else:
                        bracket_html += '<div class="bracket-group">'
                        if stage_id_br:
                            members_f = get_group_members(stage_id_br, 1)
                            if not members_f.empty:
                                for _, r in members_f.iterrows():
                                    bracket_html += f'<div class="bracket-player pending-player"><span>{r["name"]}</span><span>—</span></div>'
                            else:
                                for i in range(sd.group_size):
                                    bracket_html += f'<div class="bracket-player pending-player"><span>???</span><span>—</span></div>'
                        else:
                            for i in range(sd.group_size):
                                bracket_html += f'<div class="bracket-player pending-player"><span>???</span><span>—</span></div>'
                        bracket_html += '</div>'
                else:
                    # Финал дронов
                    fin_standings = compute_final_standings(stage_id_br)
                    if not fin_standings.empty and int(fin_standings.iloc[0].get("heats_played", 0)) > 0:
                        bracket_html += '<div class="bracket-group">'
                        medals_html = {1: "🥇", 2: "🥈", 3: "🥉"}
                        medal_class = {1: "gold", 2: "silver", 3: "bronze"}
                        for _, fr in fin_standings.iterrows():
                            rank = int(fr["rank"])
                            cls = medal_class.get(rank, "")
                            medal = medals_html.get(rank, f"{rank}.")
                            bonus = " +1б" if int(fr["bonus"]) > 0 else ""
                            bracket_html += f'<div class="bracket-player {cls}">'
                            bracket_html += f'<span>{medal} {fr["name"]}</span>'
                            bracket_html += f'<span><b>{int(fr["total"])} оч.</b> ({int(fr["wins"])} поб.{bonus})</span>'
                            bracket_html += '</div>'
                        bracket_html += '</div>'
                    else:
                        bracket_html += '<div class="bracket-group">'
                        members_f = get_group_members(stage_id_br, 1)
                        if not members_f.empty:
                            for _, r in members_f.iterrows():
                                bracket_html += f'<div class="bracket-player pending-player"><span>{r["name"]}</span><span>—</span></div>'
                        else:
                            for i in range(sd.group_size):
                                bracket_html += f'<div class="bracket-player pending-player"><span>???</span><span>—</span></div>'
                        bracket_html += '</div>'
            elif stage_id_br:
                all_groups_br = get_all_groups(stage_id_br)
                for gno in sorted(all_groups_br.keys()):
                    members = all_groups_br[gno]
                    bracket_html += f'<div class="bracket-group">'
                    bracket_html += f'<div class="bracket-group-title">Гр. {gno}</div>'
                    if is_sim:
                        # Для симулятора: показываем очки из агрегации
                        sim_rank = compute_sim_group_ranking(stage_id_br, gno, scoring_mode)
                        if not sim_rank.empty and int(sim_rank["total_points"].sum()) > 0:
                            for _, sr in sim_rank.iterrows():
                                rank = int(sr["rank"])
                                cls = "advancing" if rank <= sd.qualifiers else "eliminated"
                                bracket_html += f'<div class="bracket-player {cls}">'
                                bracket_html += f'<span>{rank}. {sr["name"]}</span>'
                                bracket_html += f'<span>{int(sr["total_points"])} оч.</span>'
                                bracket_html += '</div>'
                        elif not members.empty:
                            for i, (_, r) in enumerate(members.iterrows()):
                                bracket_html += f'<div class="bracket-player pending-player"><span>{i+1}. {r["name"]}</span><span>—</span></div>'
                    else:
                        # Для дронов: один вылет, показываем время
                        results = get_heat_results(stage_id_br, gno, 1)
                        if results:
                            for r in results:
                                place = r["place"]
                                cls = "advancing" if place <= sd.qualifiers else "eliminated"
                                time_str = format_time(r.get("time_seconds"))
                                bracket_html += f'<div class="bracket-player {cls}">'
                                bracket_html += f'<span>{place}. {r["name"]}</span>'
                                bracket_html += f'<span>{time_str}</span>'
                                bracket_html += '</div>'
                        elif not members.empty:
                            for i, (_, r) in enumerate(members.iterrows()):
                                bracket_html += f'<div class="bracket-player pending-player"><span>{i+1}. {r["name"]}</span><span>—</span></div>'
                    bracket_html += '</div>'
            else:
                for gno in range(1, sd.group_count + 1):
                    bracket_html += f'<div class="bracket-group">'
                    bracket_html += f'<div class="bracket-group-title">Гр. {gno}</div>'
                    for i in range(sd.group_size):
                        bracket_html += f'<div class="bracket-player pending-player"><span>{i+1}. ???</span><span>—</span></div>'
                    bracket_html += '</div>'

            bracket_html += '</div>'  # bracket-groups-col

            # Коннектор — внутри groups-row, растягивается на высоту групп
            if has_connector:
                bracket_html += '<div class="bracket-connector">'
                bracket_html += '<div class="bracket-conn-top"></div>'
                bracket_html += '<div class="bracket-conn-bottom"></div>'
                bracket_html += '</div>'

            bracket_html += '</div>'  # bracket-groups-row
            bracket_html += '</div>'  # bracket-round-wrapper

        bracket_html += '</div>'

        st.markdown(bracket_html, unsafe_allow_html=True)
        st.caption("🟢 Зелёный = проходит | 🔴 Красный = выбывает | 🥇🥈🥉 Медали финала")

        # --- Drag-and-drop: перемещение участников между группами ---
        if t_status == "bracket":
            active_dnd = get_active_stage(tournament_id)
            if active_dnd is not None:
                dnd_stage_id = int(active_dnd["id"])
                dnd_stage_idx = int(active_dnd["stage_idx"])
                dnd_sd = bracket[dnd_stage_idx]
                dnd_sname = dnd_sd.display_name.get(lang, dnd_sd.code)

                # Проверяем, есть ли уже результаты для этого этапа
                dnd_has_results = False
                dnd_all_groups = get_all_groups(dnd_stage_id)
                for gno in dnd_all_groups:
                    if is_sim:
                        for tr in [1, 2]:
                            for att in [1, 2, 3]:
                                if get_heat_results(dnd_stage_id, gno, att, tr):
                                    dnd_has_results = True
                                    break
                            if dnd_has_results:
                                break
                    else:
                        if get_heat_results(dnd_stage_id, gno, 1):
                            dnd_has_results = True
                    if dnd_has_results:
                        break

                if dnd_sd.code != "F" and dnd_all_groups:
                    st.divider()
                    with st.expander(f"🔀 Перемещение участников между группами ({dnd_sname})", expanded=False):
                        if dnd_has_results:
                            st.warning("🔒 Перемещение недоступно — уже есть введённые результаты для этого этапа. "
                                       "Удалите результаты или откатите этап, чтобы переместить участников.")
                        else:
                            st.caption("Перетащите участников между группами, затем нажмите «Сохранить».")

                            try:
                                from streamlit_sortables import sort_items

                                # Собираем текущее расположение и маппинг отображаемых имён → pid
                                original_containers = []
                                group_id_map = {}  # gno -> group_id в БД
                                display_to_pid = {}  # "#{sn} Имя" -> pid

                                for gno in sorted(dnd_all_groups.keys()):
                                    members = dnd_all_groups[gno]
                                    gid_df = qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=?",
                                                 (dnd_stage_id, gno))
                                    if not gid_df.empty:
                                        group_id_map[gno] = int(gid_df.iloc[0]["id"])
                                    items = []
                                    for _, r in members.iterrows():
                                        sn = f"#{int(r['start_number'])}" if pd.notna(r.get("start_number")) else ""
                                        display_name = f"{sn} {r['name']}".strip()
                                        display_to_pid[display_name] = int(r["pid"])
                                        items.append(display_name)
                                    original_containers.append({
                                        "header": f"Группа {gno}",
                                        "items": items,
                                    })

                                # Рендерим drag-and-drop
                                sorted_containers = sort_items(original_containers, multi_containers=True)

                                # Сравниваем с оригиналом
                                changed = False
                                for orig, curr in zip(original_containers, sorted_containers):
                                    if orig["items"] != curr["items"]:
                                        changed = True
                                        break

                                if changed:
                                    # Валидация размеров
                                    max_size = dnd_sd.group_size
                                    valid = True
                                    for i, container in enumerate(sorted_containers):
                                        if len(container["items"]) == 0:
                                            st.error(f"Группа {i+1} не может быть пустой!")
                                            valid = False
                                            break
                                        if len(container["items"]) > max_size:
                                            st.error(f"Группа {i+1}: максимум {max_size} участников (сейчас {len(container['items'])})!")
                                            valid = False
                                            break

                                    if valid:
                                        st.info("📝 Есть изменения в составе групп.")

                                        # Показываем что изменилось
                                        for i, (orig, curr) in enumerate(zip(original_containers, sorted_containers)):
                                            if orig["items"] != curr["items"]:
                                                added = set(curr["items"]) - set(orig["items"])
                                                removed = set(orig["items"]) - set(curr["items"])
                                                gno = i + 1
                                                if added:
                                                    st.caption(f"Группа {gno}: + {', '.join(added)}")
                                                if removed:
                                                    st.caption(f"Группа {gno}: − {', '.join(removed)}")

                                        confirm_key = "confirm_dnd_save"
                                        if not st.session_state.get(confirm_key, False):
                                            if st.button("💾 Сохранить изменения", type="primary",
                                                         use_container_width=True, key="dnd_save"):
                                                st.session_state[confirm_key] = True
                                                st.rerun()
                                        else:
                                            st.warning("⚠️ Вы уверены? Состав групп будет изменён.")
                                            dnd_c1, dnd_c2 = st.columns(2)
                                            with dnd_c1:
                                                if st.button("✅ Да, сохранить", type="primary",
                                                             use_container_width=True, key="dnd_confirm"):
                                                    group_nos = sorted(dnd_all_groups.keys())
                                                    for i, container in enumerate(sorted_containers):
                                                        gno = group_nos[i]
                                                        gid = group_id_map.get(gno)
                                                        if gid is None:
                                                            continue
                                                        exec_sql("DELETE FROM group_members WHERE group_id=?", (gid,))
                                                        for item in container["items"]:
                                                            pid = display_to_pid.get(item)
                                                            if pid is not None:
                                                                exec_sql("INSERT INTO group_members(group_id, participant_id) VALUES(?,?)",
                                                                         (gid, pid))
                                                    st.session_state[confirm_key] = False
                                                    st.success("✅ Состав групп обновлён!")
                                                    st.rerun()
                                            with dnd_c2:
                                                if st.button("❌ Отмена", use_container_width=True, key="dnd_cancel"):
                                                    st.session_state[confirm_key] = False
                                                    st.rerun()
                                else:
                                    st.success("✅ Состав групп не изменён.")

                            except ImportError:
                                st.warning("Для перемещения участников установите пакет: `pip install streamlit-sortables`")

        # Кнопка перехода
        if t_status == "bracket":
            active = get_active_stage(tournament_id)
            if active is not None:
                cur_idx = int(active["stage_idx"])
                if cur_idx + 1 < len(bracket):
                    st.divider()
                    next_sd = bracket[cur_idx + 1]
                    nname = next_sd.display_name.get(lang, next_sd.code)
                    if st.button(f"➡️ Перейти к {nname}", type="primary", use_container_width=True):
                        try:
                            advance_to_next_stage(tournament_id, bracket)
                            st.success(T("saved"))
                            st.rerun()
                        except Exception as e:
                            st.error(f"⚠️ Не удалось перейти к следующему этапу: {e}")
                elif bracket[cur_idx].code == "F":
                    # Финал завершён?
                    final_sd = bracket[cur_idx]
                    final_ok, final_msg = check_stage_results_complete(int(active["id"]), final_sd, discipline, scoring_mode)
                    if final_ok:
                        if is_sim:
                            sim_fin = compute_sim_final_standings(int(active["id"]), scoring_mode)
                            has_ties = False
                            if not sim_fin.empty:
                                pts = sim_fin["total_points"].tolist()
                                has_ties = len(pts) != len(set(pts))
                            if has_ties:
                                st.divider()
                                st.warning("⚠️ В финале есть ничья — перейдите на вкладку 'Финал' для проведения доп. вылета.")
                            else:
                                st.divider()
                                if st.button("🏆 Завершить турнир", type="primary", use_container_width=True):
                                    exec_sql("UPDATE stages SET status='done' WHERE id=?", (int(active["id"]),))
                                    exec_sql("UPDATE tournaments SET status='finished' WHERE id=?", (tournament_id,))
                                    st.success("🏆 Турнир завершён!")
                                    st.balloons()
                                    st.rerun()
                        else:
                            fin_standings = compute_final_standings(int(active["id"]))
                            fin_ties = detect_final_ties(fin_standings) if not fin_standings.empty else []
                            if fin_ties:
                                st.divider()
                                st.warning("⚠️ В финале есть ничья — перейдите на вкладку 'Финал' для проведения доп. вылета.")
                            else:
                                st.divider()
                                if st.button("🏆 Завершить турнир", type="primary", use_container_width=True):
                                    exec_sql("UPDATE stages SET status='done' WHERE id=?", (int(active["id"]),))
                                    exec_sql("UPDATE tournaments SET status='finished' WHERE id=?", (tournament_id,))
                                    st.success("🏆 Турнир завершён!")
                                    st.balloons()
                                    st.rerun()

                # --- Кнопка отката на предыдущий этап ---
                st.divider()
                cur_sd_name = bracket[cur_idx].display_name.get(lang, bracket[cur_idx].code)
                if cur_idx == 0:
                    rollback_label = "⬅️ Вернуться в квалификацию"
                else:
                    prev_name = bracket[cur_idx - 1].display_name.get(lang, bracket[cur_idx - 1].code)
                    rollback_label = f"⬅️ Вернуться к {prev_name}"

                with st.expander(rollback_label, expanded=False):
                    st.warning(f"⚠️ **Внимание!** Это удалит все результаты этапа «{cur_sd_name}» "
                               f"и вернёт турнир на предыдущий этап. Действие необратимо!")
                    rollback_key = "confirm_rollback_stage"
                    if not st.session_state.get(rollback_key, False):
                        if st.button("🔙 Откатить этап", use_container_width=True):
                            st.session_state[rollback_key] = True
                            st.rerun()
                    else:
                        st.error("Вы уверены? Все результаты текущего этапа будут удалены.")
                        rc1, rc2 = st.columns(2)
                        with rc1:
                            if st.button("✅ Да, откатить", type="primary", use_container_width=True, key="do_rollback"):
                                rollback_to_previous_stage(tournament_id, bracket)
                                st.session_state[rollback_key] = False
                                st.success("Этап откачен!")
                                st.rerun()
                        with rc2:
                            if st.button("❌ Отмена", use_container_width=True, key="cancel_rollback"):
                                st.session_state[rollback_key] = False
                                st.rerun()

        # Кнопка отката для завершённых турниров
        if t_status == "finished":
            st.divider()
            with st.expander("⬅️ Вернуть турнир в финал", expanded=False):
                st.warning("⚠️ Это снимет статус «Завершён» и вернёт турнир в финальный этап для редактирования.")
                rollback_fin_key = "confirm_rollback_finished"
                if not st.session_state.get(rollback_fin_key, False):
                    if st.button("🔙 Вернуть в финал", use_container_width=True):
                        st.session_state[rollback_fin_key] = True
                        st.rerun()
                else:
                    st.error("Вы уверены?")
                    fc1, fc2 = st.columns(2)
                    with fc1:
                        if st.button("✅ Да", type="primary", use_container_width=True, key="do_rollback_fin"):
                            rollback_to_previous_stage(tournament_id, bracket)
                            st.session_state[rollback_fin_key] = False
                            st.success("Турнир возвращён в финал!")
                            st.rerun()
                    with fc2:
                        if st.button("❌ Отмена", use_container_width=True, key="cancel_rollback_fin"):
                            st.session_state[rollback_fin_key] = False
                            st.rerun()


# ============================================================
# TAB 4: Плей-офф (ввод результатов)
# ============================================================
with tabs[4]:
    st.subheader(T("playoff_title"))

    bracket = get_bracket_for_tournament(tournament_id)
    all_stages = get_all_stages(tournament_id)

    if t_status != "bracket" or all_stages.empty:
        st.info(T("playoff_not_started"))
    else:
        # Найдём активный этап (не финал)
        active = get_active_stage(tournament_id)
        if active is None:
            st.success("Все этапы плей-офф завершены!")
        else:
            stage_id = int(active["id"])
            stage_idx = int(active["stage_idx"])
            sd = bracket[stage_idx]
            sname = sd.display_name.get(lang, sd.code)

            # Если это финал — направляем на вкладку Финал
            if sd.code == "F":
                st.info("🏆 Сейчас идёт ФИНАЛ — вводите результаты на вкладке 'Финал'")
            elif is_sim:
                # ========== СИМУЛЯТОР: 2 трассы × 3 попытки ==========
                st.success(f"🔥 Сейчас: **{sname}**")
                st.caption("📊 Сумма очков за 6 вылетов (2 трассы × 3 попытки, макс. 24 оч.)")

                all_groups = get_all_groups(stage_id)

                col_sel1, col_sel2, col_sel3 = st.columns([2, 2, 2])
                with col_sel1:
                    group_options = list(all_groups.keys())
                    group_no = st.selectbox(T("select_group"), group_options,
                                            format_func=lambda x: f"{T('group')} {x}", key="po_group") if group_options else None
                with col_sel2:
                    track_no = st.selectbox(T("track"), [1, 2],
                                            format_func=lambda x: T("track_n").format(x), key="po_track")
                with col_sel3:
                    attempt_no = st.selectbox("Попытка", [1, 2, 3],
                                              format_func=lambda x: T("attempt_n").format(x), key="po_attempt")

                # Для команд загружаем пилотов
                po_team_map = {}
                if is_team:
                    all_po_pids = []
                    for gno, gdf in all_groups.items():
                        all_po_pids.extend([int(x) for x in gdf["pid"].tolist()])
                    if all_po_pids:
                        tp_po = qdf("SELECT participant_id, pilot1_name, pilot2_name FROM team_pilots WHERE participant_id IN ({})".format(
                            ",".join(str(x) for x in all_po_pids)))
                        for _, tpr in tp_po.iterrows():
                            po_team_map[int(tpr["participant_id"])] = (tpr["pilot1_name"], tpr["pilot2_name"])

                if group_no and group_no in all_groups:
                    members = all_groups[group_no]
                    existing = get_heat_results(stage_id, group_no, attempt_no, track_no)
                    existing_map = {r["participant_id"]: r for r in existing}
                    po_dsq_pids = get_disqualified_pids(tournament_id)

                    st.divider()
                    entity_label = "команды" if is_team else "пилота"
                    st.markdown(f"### {T('group')} {group_no} — {T('track_n').format(track_no)}, {T('attempt_n').format(attempt_no)}")
                    st.caption(f"⏱️ Лимит: {time_limit} сек | 4 {entity_label}")

                    results_to_save = []
                    for _, m in members.iterrows():
                        pid = int(m["pid"])
                        pname = m["name"]
                        is_dsq = pid in po_dsq_pids
                        ex = existing_map.get(pid, {})
                        po_pilots = po_team_map.get(pid, None) if is_team else None

                        with st.container(border=True):
                            if is_dsq:
                                dsq_label = f"**{pname}** — {T('disqualified')}" if not (is_team and po_pilots) else f"**{pname}** ({po_pilots[0]}, {po_pilots[1]}) — {T('disqualified')}"
                                st.markdown(dsq_label)
                                results_to_save.append({"pid": pid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                continue
                            if is_team and po_pilots:
                                st.markdown(f"**{pname}** ({po_pilots[0]}, {po_pilots[1]})")
                            else:
                                st.markdown(f"**{pname}**")

                            if is_team:
                                # Командный: два времени + автосумма
                                p1_lbl = po_pilots[0] if po_pilots else "Пилот 1"
                                p2_lbl = po_pilots[1] if po_pilots else "Пилот 2"
                                ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                tc1, tc2, tc3 = st.columns([2, 2, 2])
                                with tc1:
                                    t1v = st.number_input(f"⏱️ {p1_lbl}", min_value=0.0, max_value=999.0,
                                                          value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                          key=f"po_t1_{group_no}_{track_no}_{attempt_no}_{pid}", format="%.3f")
                                with tc2:
                                    t2v = st.number_input(f"⏱️ {p2_lbl}", min_value=0.0, max_value=999.0,
                                                          value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                          key=f"po_t2_{group_no}_{track_no}_{attempt_no}_{pid}", format="%.3f")
                                with tc3:
                                    tval = t1v + t2v
                                    st.metric("Сумма", f"{tval:.3f}")

                                lc1, _ = st.columns([2, 4])
                                with lc1:
                                    ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                    lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                           value=ex_laps, step=0.1,
                                                           key=f"po_l_{group_no}_{track_no}_{attempt_no}_{pid}", format="%.1f")
                            else:
                                c1, c2 = st.columns([2, 2])
                                with c1:
                                    ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                    tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                           value=_safe_time_for_input(ex_time), step=0.001,
                                                           key=f"po_t_{group_no}_{track_no}_{attempt_no}_{pid}", format="%.3f")
                                with c2:
                                    ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                    lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                           value=ex_laps, step=0.1,
                                                           key=f"po_l_{group_no}_{track_no}_{attempt_no}_{pid}", format="%.1f")

                            if tval > 0:
                                results_to_save.append({
                                    "pid": pid, "time_seconds": tval,
                                    "laps_completed": lval, "completed_all_laps": lval >= total_laps
                                })

                    st.divider()
                    c1, c2 = st.columns([2, 1])
                    with c1:
                        if st.button("💾 СОХРАНИТЬ РЕЗУЛЬТАТЫ", type="primary", use_container_width=True, key="po_save_sim"):
                            if len(results_to_save) == len(members):
                                save_heat(stage_id, group_no, attempt_no, results_to_save,
                                          is_final=False, track_no=track_no, scoring=SIM_SCORING)
                                st.success(T("saved"))
                                st.balloons()
                                st.rerun()
                            else:
                                st.error("Введите результаты для всех участников!" if is_team else "Введите результаты для всех пилотов!")

                    # Таблица текущего вылета
                    results = get_heat_results(stage_id, group_no, attempt_no, track_no)
                    if results:
                        entity_col = "Команда" if is_team else "Пилот"
                        st.markdown(f"### 📊 Результаты: {T('track_n').format(track_no)}, {T('attempt_n').format(attempt_no)}")
                        tdata = [{"М": r["place"], entity_col: r["name"],
                                  "Время": format_time(r.get("time_seconds")),
                                  "Круги": r.get("laps_completed", "—"),
                                  "Очки": int(r.get("points", 0))} for r in results]
                        df_r = pd.DataFrame(tdata)
                        styled = style_standings_table(df_r, sd.qualifiers)
                        st.dataframe(styled, use_container_width=True, hide_index=True)

                    # Сводная таблица группы
                    st.divider()
                    entity_col2 = "Команда" if is_team else "Пилот"
                    st.markdown(f"### {T('sim_group_results')}: {T('group')} {group_no}")
                    sim_ranking = compute_sim_group_ranking(stage_id, group_no, scoring_mode)
                    if not sim_ranking.empty:
                        track_bests = get_sim_track_bests(stage_id, group_no)
                        pid_col = "participant_id" if "participant_id" in sim_ranking.columns else "pid"
                        sim_rows = []
                        for _, sr in sim_ranking.iterrows():
                            pid = int(sr[pid_col])
                            tb = track_bests.get(pid, {})
                            t1 = format_time(tb.get("t1")) if tb.get("t1") else "—"
                            t2 = format_time(tb.get("t2")) if tb.get("t2") else "—"
                            sim_rows.append({
                                "М": int(sr["rank"]), entity_col2: sr["name"],
                                "Трасса 1": t1, "Трасса 2": t2,
                                "Очки": int(sr["total_points"]),
                            })
                        sim_disp = pd.DataFrame(sim_rows)
                        styled_sim = style_standings_table(sim_disp, sd.qualifiers)
                        st.dataframe(styled_sim, use_container_width=True, hide_index=True)
                        st.caption("🟢 Проходит (топ-2) | 🔴 Выбывает")

                    # Прогресс заполнения
                    st.divider()
                    st.markdown("**Прогресс заполнения:**")
                    all_heats_filled = True
                    for tr in [1, 2]:
                        for att in [1, 2, 3]:
                            res = get_heat_results(stage_id, group_no, att, tr)
                            icon = "✅" if res else "⏳"
                            if not res:
                                all_heats_filled = False
                            st.caption(f"{icon} {T('track_n').format(tr)}, {T('attempt_n').format(att)}")

                    # Тайбрейк: проверяем ничьи на границе прохода
                    if all_heats_filled:
                        tied_groups = detect_sim_group_ties(stage_id, group_no, scoring_mode, sd.qualifiers)
                        if tied_groups:
                            st.divider()
                            st.error("⚠️ **Ничья на границе прохода!** Необходим дополнительный вылет.")

                            for tg in tied_groups:
                                # Показываем кто в ничьей
                                ranking_tb = compute_sim_group_ranking(stage_id, group_no, scoring_mode)
                                pid_col_tb = "participant_id" if "participant_id" in ranking_tb.columns else "pid"
                                tied_rows = ranking_tb[ranking_tb[pid_col_tb].isin(tg)]
                                tied_names = tied_rows["name"].tolist()
                                tied_pts = int(tied_rows.iloc[0]["total_points"]) if not tied_rows.empty else 0
                                st.warning(f"🤝 Ничья ({tied_pts} оч.): **{', '.join(tied_names)}**")

                            all_tied_pids = []
                            for tg in tied_groups:
                                all_tied_pids.extend(tg)

                            st.markdown("### 🔄 Дополнительный вылет")
                            tb_entity = "команды" if is_team else "пилоты"
                            st.caption(f"Участвуют {tb_entity} с одинаковыми очками. Результат определит кто проходит дальше.")

                            # track_no=99 — специальный маркер тайбрейка
                            existing_tb = get_heat_results(stage_id, group_no, 1, track_no=99)
                            existing_tb_map = {r["participant_id"]: r for r in existing_tb}
                            tb_dsq_pids = get_disqualified_pids(tournament_id)

                            tb_results = []
                            members_tb = get_group_members(stage_id, group_no)
                            for tpid in all_tied_pids:
                                prow = members_tb[members_tb["pid"] == tpid]
                                if prow.empty:
                                    continue
                                pname = prow.iloc[0]["name"]
                                is_dsq_tb = tpid in tb_dsq_pids
                                ex = existing_tb_map.get(tpid, {})
                                tb_pilots = po_team_map.get(tpid, None) if is_team else None

                                with st.container(border=True):
                                    if is_dsq_tb:
                                        dsq_lbl = f"**{pname}** — {T('disqualified')}" if not (is_team and tb_pilots) else f"**{pname}** ({tb_pilots[0]}, {tb_pilots[1]}) — {T('disqualified')}"
                                        st.markdown(dsq_lbl)
                                        tb_results.append({"pid": tpid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                        continue
                                    if is_team and tb_pilots:
                                        st.markdown(f"**{pname}** ({tb_pilots[0]}, {tb_pilots[1]})")
                                    else:
                                        st.markdown(f"**{pname}**")

                                    if is_team:
                                        p1l = tb_pilots[0] if tb_pilots else "Пилот 1"
                                        p2l = tb_pilots[1] if tb_pilots else "Пилот 2"
                                        ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                        tbc1, tbc2, tbc3 = st.columns([2, 2, 2])
                                        with tbc1:
                                            tb1v = st.number_input(f"⏱️ {p1l}", min_value=0.0, max_value=999.0,
                                                                   value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                                   key=f"tb_t1_{group_no}_{tpid}", format="%.3f")
                                        with tbc2:
                                            tb2v = st.number_input(f"⏱️ {p2l}", min_value=0.0, max_value=999.0,
                                                                   value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                                   key=f"tb_t2_{group_no}_{tpid}", format="%.3f")
                                        with tbc3:
                                            tval = tb1v + tb2v
                                            st.metric("Сумма", f"{tval:.3f}")
                                        tblc1, _ = st.columns([2, 4])
                                        with tblc1:
                                            ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                            lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                                   value=ex_laps, step=0.1,
                                                                   key=f"tb_l_{group_no}_{tpid}", format="%.1f")
                                    else:
                                        c1, c2 = st.columns([2, 2])
                                        with c1:
                                            ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                            tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                                   value=_safe_time_for_input(ex_time), step=0.001,
                                                                   key=f"tb_t_{group_no}_{tpid}", format="%.3f")
                                        with c2:
                                            ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                            lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                                   value=ex_laps, step=0.1,
                                                                   key=f"tb_l_{group_no}_{tpid}", format="%.1f")

                                if tval > 0:
                                    tb_results.append({
                                        "pid": tpid, "time_seconds": tval,
                                        "laps_completed": lval, "completed_all_laps": lval >= total_laps
                                    })

                            if st.button("💾 Сохранить доп. вылет", type="primary",
                                         use_container_width=True, key=f"tb_save_{group_no}"):
                                if len(tb_results) == len(all_tied_pids):
                                    save_heat(stage_id, group_no, 1, tb_results,
                                              is_final=False, track_no=99, scoring=SIM_SCORING)
                                    st.success("Сохранено! Ничья разрешена.")
                                    st.rerun()
                                else:
                                    st.error("Введите результаты для всех участников тайбрейка!")

                            if existing_tb:
                                ent_tb = "Команда" if is_team else "Пилот"
                                st.markdown("**Результаты доп. вылета:**")
                                tdata = [{"М": r["place"], ent_tb: r["name"],
                                          "Время": format_time(r.get("time_seconds")),
                                          "Круги": r.get("laps_completed", "—")} for r in existing_tb]
                                st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)

                                # Обновлённая сводка с учётом тайбрейка
                                st.markdown("### 📊 Обновлённая сводка (с учётом тайбрейка)")
                                resolved_ranking = resolve_sim_tiebreaker(stage_id, group_no, scoring_mode)
                                if not resolved_ranking.empty:
                                    pid_col_r = "participant_id" if "participant_id" in resolved_ranking.columns else "pid"
                                    track_bests_r = get_sim_track_bests(stage_id, group_no)
                                    rr_rows = []
                                    for _, sr in resolved_ranking.iterrows():
                                        pid = int(sr[pid_col_r])
                                        tbr = track_bests_r.get(pid, {})
                                        t1 = format_time(tbr.get("t1")) if tbr.get("t1") else "—"
                                        t2 = format_time(tbr.get("t2")) if tbr.get("t2") else "—"
                                        rr_rows.append({
                                            "М": int(sr["rank"]), ent_tb: sr["name"],
                                            "Трасса 1": t1, "Трасса 2": t2,
                                            "Очки": int(sr["total_points"]),
                                        })
                                    rr_disp = pd.DataFrame(rr_rows)
                                    styled_rr = style_standings_table(rr_disp, sd.qualifiers)
                                    st.dataframe(styled_rr, use_container_width=True, hide_index=True)

            else:
                # ========== ДРОНЫ: один вылет ==========
                st.success(f"🔥 Сейчас: **{sname}**")

                all_groups = get_all_groups(stage_id)

                col1, col2 = st.columns([2, 3])
                with col1:
                    group_options = list(all_groups.keys())
                    if group_options:
                        group_no = st.selectbox(T("select_group"), group_options,
                                                format_func=lambda x: f"{T('group')} {x}", key="po_group")
                    else:
                        group_no = None

                if group_no and group_no in all_groups:
                    members = all_groups[group_no]
                    existing = get_heat_results(stage_id, group_no, 1)
                    existing_map = {r["participant_id"]: r for r in existing}
                    po_dsq_pids = get_disqualified_pids(tournament_id)

                    st.divider()
                    st.markdown(f"### {T('group')} {group_no} — Вылет")
                    st.caption(f"⏱️ Лимит: {time_limit} сек | 4 пилота, проходят 2 лучших")

                    results_to_save = []
                    for _, m in members.iterrows():
                        pid = int(m["pid"])
                        pname = m["name"]
                        is_dsq = pid in po_dsq_pids
                        ex = existing_map.get(pid, {})

                        with st.container(border=True):
                            if is_dsq:
                                st.markdown(f"**{pname}** — {T('disqualified')}")
                                results_to_save.append({"pid": pid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                continue
                            st.markdown(f"**{pname}**")
                            c1, c2, c3, c4 = st.columns([2, 2, 1, 2])
                            with c1:
                                ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                       value=_safe_time_for_input(ex_time), step=0.001, key=f"po_t_{group_no}_{pid}", format="%.3f")
                            with c2:
                                ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                       value=ex_laps, step=0.1, key=f"po_l_{group_no}_{pid}", format="%.1f")
                            with c3:
                                ex_all = bool(ex.get("completed_all_laps", 0))
                                aval = st.checkbox("Все круги", value=ex_all, key=f"po_a_{group_no}_{pid}",
                                                   help="Отметьте, если пилот прошёл все круги за отведённое время")
                            with c4:
                                if tval > 0 and lval > 0:
                                    proj = tval if aval else calc_projected_time(tval, lval, total_laps)
                                    st.metric("Расчётное", format_time(proj))

                            if tval > 0:
                                results_to_save.append({
                                    "pid": pid, "time_seconds": tval,
                                    "laps_completed": lval, "completed_all_laps": aval
                                })

                    st.divider()
                    c1, c2 = st.columns([2, 1])
                    with c1:
                        if st.button("💾 СОХРАНИТЬ РЕЗУЛЬТАТЫ", type="primary", use_container_width=True, key="po_save_drone"):
                            if len(results_to_save) == len(members):
                                save_heat(stage_id, group_no, 1, results_to_save, is_final=False)
                                st.success(T("saved"))
                                st.balloons()
                                st.rerun()
                            else:
                                st.error("Введите результаты для всех пилотов!")

                    # Таблица результатов группы
                    results = get_heat_results(stage_id, group_no, 1)
                    if results:
                        st.markdown("### 📊 Результаты")
                        tdata = []
                        for r in results:
                            tdata.append({
                                "М": r["place"], "Пилот": r["name"],
                                "Время": format_time(r.get("time_seconds")),
                                "Круги": r.get("laps_completed", "—"),
                                "Все": "✅" if r.get("completed_all_laps") else "—",
                                "Расч.": format_time(r.get("projected_time")),
                            })
                        df_r = pd.DataFrame(tdata)
                        styled = style_standings_table(df_r, sd.qualifiers)
                        st.dataframe(styled, use_container_width=True, hide_index=True)
                        st.caption("🟢 Проходит | 🔴 Выбывает")

# ============================================================
# TAB 5: Финал
# ============================================================
with tabs[5]:
    st.subheader(f"🏆 {T('final_title')}")

    bracket = get_bracket_for_tournament(tournament_id)
    all_stages = get_all_stages(tournament_id)

    # Найдём финальный этап
    final_stage = None
    if not all_stages.empty:
        for idx, sd in enumerate(bracket):
            if sd.code == "F":
                row = all_stages[all_stages["stage_idx"] == idx]
                if not row.empty:
                    final_stage = row.iloc[0]
                break

    if final_stage is None:
        st.info("Финал ещё не начался. Завершите предыдущие этапы.")
    else:
        stage_id = int(final_stage["id"])
        members = get_group_members(stage_id, 1)
        is_finished = t_status == "finished"

        if members.empty:
            st.warning("Финалисты не определены")
        elif is_sim:
            # ============================================================
            # ФИНАЛ ДЛЯ СИМУЛЯТОРА: 2 трассы × 3 попытки
            # ============================================================
            if is_finished:
                st.success("🏆 **ТУРНИР ЗАВЕРШЁН!**")
            else:
                st.success(f"🏆 Финалисты: {', '.join(members['name'].tolist())}")

            st.caption("📊 Сумма очков за 6 вылетов (2 трассы × 3 попытки, макс. 24 оч.)")

            # Для команд загружаем пилотов
            fn_team_map = {}
            if is_team:
                fn_pids = [int(x) for x in members["pid"].tolist()]
                if fn_pids:
                    tp_fn = qdf("SELECT participant_id, pilot1_name, pilot2_name FROM team_pilots WHERE participant_id IN ({})".format(
                        ",".join(str(x) for x in fn_pids)))
                    for _, tpr in tp_fn.iterrows():
                        fn_team_map[int(tpr["participant_id"])] = (tpr["pilot1_name"], tpr["pilot2_name"])

            # Ввод результатов по трасса/попытка
            if not is_finished:
                st.markdown("### Ввод результатов финала")
                col_sel1, col_sel2 = st.columns([2, 2])
                with col_sel1:
                    fn_track = st.selectbox(T("track"), [1, 2],
                                            format_func=lambda x: T("track_n").format(x), key="fn_track")
                with col_sel2:
                    fn_attempt = st.selectbox("Попытка", [1, 2, 3],
                                              format_func=lambda x: T("attempt_n").format(x), key="fn_attempt")

                existing = get_heat_results(stage_id, 1, fn_attempt, fn_track)
                existing_map = {r["participant_id"]: r for r in existing}
                fn_dsq_pids = get_disqualified_pids(tournament_id)

                st.divider()
                fn_entity = "команды" if is_team else "пилота"
                st.markdown(f"### {T('track_n').format(fn_track)}, {T('attempt_n').format(fn_attempt)}")
                st.caption(f"⏱️ Лимит: {time_limit} сек | 4 {fn_entity}")

                results_to_save = []
                for _, m in members.iterrows():
                    pid = int(m["pid"])
                    pname = m["name"]
                    is_dsq = pid in fn_dsq_pids
                    ex = existing_map.get(pid, {})
                    fn_pilots = fn_team_map.get(pid, None) if is_team else None

                    with st.container(border=True):
                        if is_dsq:
                            dsq_label = f"**{pname}** — {T('disqualified')}" if not (is_team and fn_pilots) else f"**{pname}** ({fn_pilots[0]}, {fn_pilots[1]}) — {T('disqualified')}"
                            st.markdown(dsq_label)
                            results_to_save.append({"pid": pid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                            continue
                        if is_team and fn_pilots:
                            st.markdown(f"**{pname}** ({fn_pilots[0]}, {fn_pilots[1]})")
                        else:
                            st.markdown(f"**{pname}**")

                        if is_team:
                            p1_lbl = fn_pilots[0] if fn_pilots else "Пилот 1"
                            p2_lbl = fn_pilots[1] if fn_pilots else "Пилот 2"
                            ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                            fc1, fc2, fc3 = st.columns([2, 2, 2])
                            with fc1:
                                ft1v = st.number_input(f"⏱️ {p1_lbl}", min_value=0.0, max_value=999.0,
                                                       value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                       key=f"fn_t1_{fn_track}_{fn_attempt}_{pid}", format="%.3f")
                            with fc2:
                                ft2v = st.number_input(f"⏱️ {p2_lbl}", min_value=0.0, max_value=999.0,
                                                       value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                       key=f"fn_t2_{fn_track}_{fn_attempt}_{pid}", format="%.3f")
                            with fc3:
                                tval = ft1v + ft2v
                                st.metric("Сумма", f"{tval:.3f}")
                            flc1, _ = st.columns([2, 4])
                            with flc1:
                                ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                       value=ex_laps, step=0.1,
                                                       key=f"fn_l_{fn_track}_{fn_attempt}_{pid}", format="%.1f")
                        else:
                            c1, c2 = st.columns([2, 2])
                            with c1:
                                ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                       value=_safe_time_for_input(ex_time), step=0.001,
                                                       key=f"fn_t_{fn_track}_{fn_attempt}_{pid}", format="%.3f")
                            with c2:
                                ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                       value=ex_laps, step=0.1,
                                                       key=f"fn_l_{fn_track}_{fn_attempt}_{pid}", format="%.1f")

                    if tval > 0:
                        results_to_save.append({
                            "pid": pid, "time_seconds": tval,
                            "laps_completed": lval, "completed_all_laps": lval >= total_laps
                        })

                if st.button("💾 СОХРАНИТЬ РЕЗУЛЬТАТЫ", type="primary", use_container_width=True, key="fn_save"):
                    if len(results_to_save) == len(members):
                        save_heat(stage_id, 1, fn_attempt, results_to_save,
                                  is_final=False, track_no=fn_track, scoring=SIM_SCORING)
                        st.success(T("saved"))
                        st.rerun()
                    else:
                        st.error("Введите результаты для всех!")

                # Результаты текущего вылета
                fn_ent_col = "Команда" if is_team else "Пилот"
                results = get_heat_results(stage_id, 1, fn_attempt, fn_track)
                if results:
                    tdata = [{"М": r["place"], fn_ent_col: r["name"],
                              "Время": format_time(r.get("time_seconds")),
                              "Круги": r.get("laps_completed", "—"),
                              "Очки": int(r.get("points", 0))} for r in results]
                    st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)

            # Прогресс заполнения
            st.divider()
            st.markdown("**Прогресс заполнения финала:**")
            all_filled = True
            for tr in [1, 2]:
                for att in [1, 2, 3]:
                    res = get_heat_results(stage_id, 1, att, tr)
                    icon = "✅" if res else "⏳"
                    if not res:
                        all_filled = False
                    st.caption(f"{icon} {T('track_n').format(tr)}, {T('attempt_n').format(att)}")

            # Итоговая таблица
            st.divider()
            st.markdown(f"### 🏆 {T('final_standings')}")

            fn_standings_col = "Команда" if is_team else "Пилот"
            sim_standings = compute_sim_final_standings(stage_id, scoring_mode)
            if not sim_standings.empty:
                track_bests_fin = get_sim_track_bests(stage_id, 1)
                pid_col_fin = "participant_id" if "participant_id" in sim_standings.columns else "pid"
                medal_data = []
                for _, row in sim_standings.iterrows():
                    rank = int(row["rank"])
                    pid = int(row[pid_col_fin])
                    tb = track_bests_fin.get(pid, {})
                    t1 = format_time(tb.get("t1")) if tb.get("t1") else "—"
                    t2 = format_time(tb.get("t2")) if tb.get("t2") else "—"
                    medals = {1: "🥇", 2: "🥈", 3: "🥉"}
                    medal_data.append({
                        "М": rank,
                        "": medals.get(rank, ""),
                        fn_standings_col: row["name"],
                        "Трасса 1": t1,
                        "Трасса 2": t2,
                        "Очки": int(row["total_points"]),
                    })
                df_final = pd.DataFrame(medal_data)
                styled_final = style_final_podium(df_final)
                st.dataframe(styled_final, use_container_width=True, hide_index=True)

                if is_finished:
                    champion = sim_standings.iloc[0]["name"]
                    st.success(f"🏆 **ЧЕМПИОН: {champion}!** {T('champion')}")
                else:
                    # Проверяем ничьи
                    if all_filled:
                        # Detect ties for sim
                        pts_col = "total_points"
                        pts_vals = sim_standings[pts_col].tolist()
                        sim_tied_groups = []
                        seen = set()
                        for i, p in enumerate(pts_vals):
                            if i in seen:
                                continue
                            group = [sim_standings.iloc[i]["participant_id"] if "participant_id" in sim_standings.columns
                                     else sim_standings.iloc[i].get("pid", i)]
                            for j in range(i + 1, len(pts_vals)):
                                if pts_vals[j] == p:
                                    pid_j = sim_standings.iloc[j]["participant_id"] if "participant_id" in sim_standings.columns \
                                        else sim_standings.iloc[j].get("pid", j)
                                    group.append(pid_j)
                                    seen.add(j)
                            if len(group) > 1:
                                sim_tied_groups.append([int(x) for x in group])

                        if sim_tied_groups:
                            st.divider()
                            st.error("⚠️ **Обнаружена ничья!** Необходим дополнительный вылет для определения мест.")

                            for tg in sim_tied_groups:
                                pid_col = "participant_id" if "participant_id" in sim_standings.columns else "pid"
                                tied_rows_fn = sim_standings[sim_standings[pid_col].isin(tg)]
                                tied_names = tied_rows_fn["name"].tolist()
                                tied_pts = int(tied_rows_fn.iloc[0]["total_points"]) if not tied_rows_fn.empty else 0
                                st.warning(f"🤝 Ничья ({tied_pts} оч.): **{', '.join(tied_names)}**")

                            # Тайбрейк: дополнительный вылет
                            all_tied_pids = []
                            for tg in sim_tied_groups:
                                all_tied_pids.extend(tg)

                            group_id = int(qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=1",
                                               (stage_id,)).iloc[0]["id"])
                            max_heat_df = qdf("SELECT MAX(heat_no) as mx FROM heats WHERE group_id=? AND track_no=1",
                                              (group_id,))
                            max_heat = int(max_heat_df.iloc[0]["mx"]) if not max_heat_df.empty and max_heat_df.iloc[0]["mx"] is not None else 3
                            next_tb = max_heat + 1

                            fn_tb_entity = "команды" if is_team else "пилоты"
                            st.markdown(f"### 🔄 Дополнительный вылет #{next_tb - 3}")
                            st.caption(f"Участвуют {fn_tb_entity} с одинаковыми очками. Результат определит места.")

                            existing_tb = get_heat_results(stage_id, 1, next_tb, track_no=1)
                            existing_tb_map = {r["participant_id"]: r for r in existing_tb}
                            fn_tb_dsq_pids = get_disqualified_pids(tournament_id)

                            tb_results = []
                            pid_col = "participant_id" if "participant_id" in sim_standings.columns else "pid"
                            for tpid in all_tied_pids:
                                prow_df = sim_standings[sim_standings[pid_col] == tpid]
                                if prow_df.empty:
                                    continue
                                prow = prow_df.iloc[0]
                                pname = prow["name"]
                                is_dsq_tb = tpid in fn_tb_dsq_pids
                                ex = existing_tb_map.get(tpid, {})
                                fn_tb_pilots = fn_team_map.get(tpid, None) if is_team else None

                                with st.container(border=True):
                                    if is_dsq_tb:
                                        dsq_lbl = f"**{pname}** — {T('disqualified')}" if not (is_team and fn_tb_pilots) else f"**{pname}** ({fn_tb_pilots[0]}, {fn_tb_pilots[1]}) — {T('disqualified')}"
                                        st.markdown(dsq_lbl)
                                        tb_results.append({"pid": tpid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                        continue
                                    if is_team and fn_tb_pilots:
                                        st.markdown(f"**{pname}** ({fn_tb_pilots[0]}, {fn_tb_pilots[1]})")
                                    else:
                                        st.markdown(f"**{pname}**")

                                    if is_team:
                                        p1l = fn_tb_pilots[0] if fn_tb_pilots else "Пилот 1"
                                        p2l = fn_tb_pilots[1] if fn_tb_pilots else "Пилот 2"
                                        ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                        fntbc1, fntbc2, fntbc3 = st.columns([2, 2, 2])
                                        with fntbc1:
                                            fntb1v = st.number_input(f"⏱️ {p1l}", min_value=0.0, max_value=999.0,
                                                                     value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                                     key=f"fn_tb_t1_{next_tb}_{tpid}", format="%.3f")
                                        with fntbc2:
                                            fntb2v = st.number_input(f"⏱️ {p2l}", min_value=0.0, max_value=999.0,
                                                                     value=_safe_time_for_input(ex_time) / 2, step=0.001,
                                                                     key=f"fn_tb_t2_{next_tb}_{tpid}", format="%.3f")
                                        with fntbc3:
                                            tval = fntb1v + fntb2v
                                            st.metric("Сумма", f"{tval:.3f}")
                                        fntblc, _ = st.columns([2, 4])
                                        with fntblc:
                                            ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                            lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                                   value=ex_laps, step=0.1,
                                                                   key=f"fn_tb_l_{next_tb}_{tpid}", format="%.1f")
                                    else:
                                        c1, c2 = st.columns([2, 2])
                                        with c1:
                                            ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                            tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                                   value=_safe_time_for_input(ex_time), step=0.001,
                                                                   key=f"tb_t_{next_tb}_{tpid}", format="%.3f")
                                        with c2:
                                            ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                            lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                                   value=ex_laps, step=0.1,
                                                                   key=f"tb_l_{next_tb}_{tpid}", format="%.1f")

                                if tval > 0:
                                    tb_results.append({
                                        "pid": tpid, "time_seconds": tval,
                                        "laps_completed": lval, "completed_all_laps": lval >= total_laps
                                    })

                            if st.button(f"💾 Сохранить доп. вылет #{next_tb - 3}", type="primary",
                                         use_container_width=True, key=f"tb_save_{next_tb}"):
                                if len(tb_results) == len(all_tied_pids):
                                    save_heat(stage_id, 1, next_tb, tb_results,
                                              is_final=False, track_no=1, scoring=SIM_SCORING)
                                    st.success("Сохранено! Проверьте итоговую таблицу.")
                                    st.rerun()
                                else:
                                    st.error("Введите результаты для всех участников тайбрейка!")

                        else:
                            # Нет ничьих — можно завершить
                            st.divider()
                            if st.button("🏆 Завершить турнир", type="primary", use_container_width=True,
                                         key="finish_tournament"):
                                exec_sql("UPDATE stages SET status='done' WHERE id=?", (stage_id,))
                                exec_sql("UPDATE tournaments SET status='finished' WHERE id=?", (tournament_id,))
                                st.success("🏆 Турнир завершён!")
                                st.balloons()
                                st.rerun()
                    else:
                        st.info("Заполните все 6 вылетов (2 трассы × 3 попытки) для определения итогов")
            else:
                st.info("Введите результаты финала")

        else:
            # ============================================================
            # ФИНАЛ ДЛЯ ДРОНОВ: 3 вылета + бонус
            # ============================================================
            if is_finished:
                st.success("🏆 **ТУРНИР ЗАВЕРШЁН!**")
            else:
                st.success(f"🏆 Финалисты: {', '.join(members['name'].tolist())}")
            st.caption(T("bonus_note"))

            # 3 вылета
            for heat_no in range(1, 4):
                st.divider()
                st.markdown(f"### {T('heat_n').format(heat_no)}")

                existing = get_heat_results(stage_id, 1, heat_no)
                existing_map = {r["participant_id"]: r for r in existing}
                fn_dsq_pids = get_disqualified_pids(tournament_id)

                if is_finished:
                    if existing:
                        tdata = [{"М": r["place"], "Пилот": r["name"],
                                  "Время": format_time(r["time_seconds"]),
                                  "Круги": r.get("laps_completed", "—"),
                                  "Все": "✅" if r.get("completed_all_laps") else "—",
                                  "Очки": f"+{r['points']}"} for r in existing]
                        st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)
                    else:
                        st.caption("Нет результатов")
                else:
                    results_to_save = []
                    for _, m in members.iterrows():
                        pid = int(m["pid"])
                        pname = m["name"]
                        is_dsq = pid in fn_dsq_pids
                        ex = existing_map.get(pid, {})

                        with st.container(border=True):
                            if is_dsq:
                                st.markdown(f"**{pname}** — {T('disqualified')}")
                                results_to_save.append({"pid": pid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                continue
                            st.markdown(f"**{pname}**")
                            c1, c2, c3, c4 = st.columns([2, 2, 1, 2])
                            with c1:
                                ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                       value=_safe_time_for_input(ex_time), step=0.001,
                                                       key=f"fn_t_{heat_no}_{pid}", format="%.3f")
                            with c2:
                                ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                       value=ex_laps, step=0.1,
                                                       key=f"fn_l_{heat_no}_{pid}", format="%.1f")
                            with c3:
                                ex_all = bool(ex.get("completed_all_laps", 0))
                                aval = st.checkbox("Все круги", value=ex_all, key=f"fn_a_{heat_no}_{pid}",
                                                   help="Отметьте, если пилот прошёл все круги за отведённое время")
                            with c4:
                                if tval > 0 and lval > 0:
                                    proj = tval if aval else calc_projected_time(tval, laps_completed=lval, total_laps=total_laps)
                                    st.metric("Расчётное", format_time(proj))

                        if tval > 0:
                            results_to_save.append({
                                "pid": pid, "time_seconds": tval,
                                "laps_completed": lval, "completed_all_laps": aval
                            })

                    if st.button(f"💾 Сохранить вылет {heat_no}", type="primary", key=f"fn_save_{heat_no}"):
                        if len(results_to_save) == len(members):
                            save_heat(stage_id, 1, heat_no, results_to_save, is_final=True)
                            st.success(T("saved"))
                            st.rerun()
                        else:
                            st.error("Введите результаты для всех!")

                    results = get_heat_results(stage_id, 1, heat_no)
                    if results:
                        tdata = [{"М": r["place"], "Пилот": r["name"],
                                  "Время": format_time(r["time_seconds"]),
                                  "Очки": f"+{r['points']}"} for r in results]
                        st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)

            # Итоговая таблица
            st.divider()
            st.markdown(f"### 🏆 {T('final_standings')}")

            standings = compute_final_standings(stage_id)
            if not standings.empty:
                medal_data = []
                for _, row in standings.iterrows():
                    rank = int(row["rank"])
                    medals = {1: "🥇", 2: "🥈", 3: "🥉"}
                    bonus_str = "+1" if int(row["bonus"]) > 0 else ""
                    medal_data.append({
                        "М": rank,
                        "": medals.get(rank, ""),
                        "Пилот": row["name"],
                        "Очки": int(row["total"]),
                        "Баллы": int(row["total_points"]),
                        "Бонус": bonus_str,
                        "Побед": int(row["wins"]),
                    })
                df_final = pd.DataFrame(medal_data)
                styled_final = style_final_podium(df_final)
                st.dataframe(styled_final, use_container_width=True, hide_index=True)

                if is_finished:
                    champion = standings.iloc[0]["name"]
                    st.success(f"🏆 **ЧЕМПИОН: {champion}!** {T('champion')}")
                else:
                    tied_groups = detect_final_ties(standings)
                    has_basic_3 = int(standings.iloc[0].get("heats_played", 0)) >= 3

                    if has_basic_3 and tied_groups:
                        st.divider()
                        st.error("⚠️ **Обнаружена ничья!** Необходим дополнительный вылет для определения мест.")

                        group_id = int(qdf("SELECT id FROM groups WHERE stage_id=? AND group_no=1",
                                           (stage_id,)).iloc[0]["id"])
                        max_heat_df = qdf("SELECT MAX(heat_no) as mx FROM heats WHERE group_id=?", (group_id,))
                        max_heat = int(max_heat_df.iloc[0]["mx"]) if not max_heat_df.empty and max_heat_df.iloc[0]["mx"] is not None else 3
                        next_tb = max_heat + 1

                        for tg in tied_groups:
                            tied_rows_dr = standings[standings["pid"].isin(tg)]
                            tied_names = tied_rows_dr["name"].tolist()
                            tied_total = int(tied_rows_dr.iloc[0]["total"]) if not tied_rows_dr.empty else 0
                            st.warning(f"🤝 Ничья ({tied_total} оч.): **{', '.join(tied_names)}**")

                        all_tied_pids = []
                        for tg in tied_groups:
                            all_tied_pids.extend(tg)

                        st.markdown(f"### 🔄 Дополнительный вылет #{next_tb - 3}")
                        st.caption("Участвуют только пилоты с одинаковым количеством очков. Результат определит итоговые места.")

                        existing_tb = get_heat_results(stage_id, 1, next_tb)
                        existing_tb_map = {r["participant_id"]: r for r in existing_tb}
                        fn_dr_tb_dsq_pids = get_disqualified_pids(tournament_id)

                        tb_results = []
                        for tpid in all_tied_pids:
                            prow_df = standings[standings["pid"] == tpid]
                            if prow_df.empty:
                                continue
                            prow = prow_df.iloc[0]
                            pname = prow["name"]
                            is_dsq_tb = tpid in fn_dr_tb_dsq_pids
                            ex = existing_tb_map.get(tpid, {})

                            with st.container(border=True):
                                if is_dsq_tb:
                                    st.markdown(f"**{pname}** — {T('disqualified')}")
                                    tb_results.append({"pid": tpid, "time_seconds": 9999.0, "laps_completed": 0.0, "completed_all_laps": False})
                                    continue
                                st.markdown(f"**{pname}**")
                                c1, c2, c3, c4 = st.columns([2, 2, 1, 2])
                                with c1:
                                    ex_time = float(ex["time_seconds"]) if ex.get("time_seconds") else 0.0
                                    tval = st.number_input("Время (сек)", min_value=0.0, max_value=999.0,
                                                           value=_safe_time_for_input(ex_time), step=0.001,
                                                           key=f"tb_t_{next_tb}_{tpid}", format="%.3f")
                                with c2:
                                    ex_laps = float(ex["laps_completed"]) if ex.get("laps_completed") else 0.0
                                    lval = st.number_input("Круги.Препятствия", min_value=0.0, max_value=99.0,
                                                           value=ex_laps, step=0.1,
                                                           key=f"tb_l_{next_tb}_{tpid}", format="%.1f")
                                with c3:
                                    ex_all = bool(ex.get("completed_all_laps", 0))
                                    aval = st.checkbox("Все круги", value=ex_all, key=f"tb_a_{next_tb}_{tpid}",
                                                       help="Отметьте, если пилот прошёл все круги за отведённое время")
                                with c4:
                                    if tval > 0 and lval > 0:
                                        proj = tval if aval else calc_projected_time(tval, lval, total_laps)
                                        st.metric("Расчётное", format_time(proj))

                            if tval > 0:
                                tb_results.append({
                                    "pid": tpid, "time_seconds": tval,
                                    "laps_completed": lval, "completed_all_laps": aval
                                })

                        if st.button(f"💾 Сохранить доп. вылет #{next_tb - 3}", type="primary",
                                     use_container_width=True, key=f"tb_save_{next_tb}"):
                            if len(tb_results) == len(all_tied_pids):
                                save_heat(stage_id, 1, next_tb, tb_results, is_final=False)
                                st.success("Сохранено! Проверьте итоговую таблицу.")
                                st.rerun()
                            else:
                                st.error("Введите результаты для всех участников тайбрейка!")

                        if existing_tb:
                            st.markdown(f"**Результаты доп. вылета #{next_tb - 3}:**")
                            tdata = [{"М": r["place"], "Пилот": r["name"],
                                      "Время": format_time(r.get("time_seconds")),
                                      "Круги": r.get("laps_completed", "—")} for r in existing_tb]
                            st.dataframe(pd.DataFrame(tdata), use_container_width=True, hide_index=True)

                    elif has_basic_3 and not tied_groups:
                        st.divider()
                        if st.button("🏆 Завершить турнир", type="primary", use_container_width=True, key="finish_tournament"):
                            exec_sql("UPDATE stages SET status='done' WHERE id=?", (stage_id,))
                            exec_sql("UPDATE tournaments SET status='finished' WHERE id=?", (tournament_id,))
                            st.success("🏆 Турнир завершён!")
                            st.balloons()
                            st.rerun()
            else:
                st.info("Введите результаты вылетов выше")

# ============================================================
# TAB 6: Результаты (итоговая таблица всех участников)
# ============================================================
with tabs[6]:
    st.subheader("📋 Итоговые результаты турнира")

    if t_status != "finished":
        st.info("Итоговая таблица будет доступна после завершения турнира.")
    else:
        overall_df = compute_overall_standings(tournament_id)
        if overall_df.empty:
            st.warning("Нет данных для отображения.")
        else:
            entity_label = "Команда" if is_team else "Пилот"

            st.success(f"🏆 Турнир завершён! Всего участников: {len(overall_df)}")

            top3 = overall_df[overall_df["place"] <= 3]
            if len(top3) >= 3:
                pc1, pc2, pc3 = st.columns(3)
                medals = {1: ("🥇", "gold"), 2: ("🥈", "silver"), 3: ("🥉", "bronze")}
                for col, place_num in zip([pc2, pc1, pc3], [1, 2, 3]):
                    row = top3[top3["place"] == place_num]
                    if not row.empty:
                        r = row.iloc[0]
                        emoji, _ = medals[place_num]
                        with col:
                            st.markdown(f"### {emoji} {place_num} место")
                            st.markdown(f"**{r['name']}**")
                            st.caption(r["detail"])

            st.divider()
            st.markdown("### Полная таблица")

            display_rows = []
            for _, row in overall_df.iterrows():
                medals_map = {1: "🥇", 2: "🥈", 3: "🥉"}
                display_rows.append({
                    "Место": int(row["place"]),
                    "": medals_map.get(int(row["place"]), ""),
                    entity_label: row["name"],
                    "Этап выбывания": row["stage"],
                    "Результат": row["detail"],
                })
            df_display = pd.DataFrame(display_rows)

            def style_overall(row):
                place = row["Место"]
                if place == 1:
                    return ["background-color: rgba(255, 215, 0, 0.15)"] * len(row)
                elif place == 2:
                    return ["background-color: rgba(192, 192, 192, 0.18)"] * len(row)
                elif place == 3:
                    return ["background-color: rgba(205, 127, 50, 0.15)"] * len(row)
                elif place == 4:
                    return ["background-color: rgba(150, 150, 150, 0.08)"] * len(row)
                return [""] * len(row)

            styled = df_display.style.apply(style_overall, axis=1)
            st.dataframe(styled, use_container_width=True, hide_index=True)

            # Экспорт таблицы
            csv_data = df_display.to_csv(index=False).encode("utf-8-sig")
            st.download_button("📥 Скачать таблицу результатов (CSV)", data=csv_data,
                               file_name=f"results_{tournament_id}.csv", mime="text/csv")
